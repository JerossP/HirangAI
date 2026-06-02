"""
app.py
──────
HirangAI — Streamlit frontend entry point.

Run with:
    streamlit run app.py
"""

from __future__ import annotations

import sys
import logging
from dotenv import load_dotenv
load_dotenv(override=True)  # Load environment variables with force override!

import streamlit as st
from resume_parser import parse_resume
from candidate_evaluator import evaluate_candidate

# Setup terminal-only logger
logger = logging.getLogger("HirangAI.pipeline")
if not logger.handlers:
    _ch = logging.StreamHandler(sys.stderr)
    _ch.setFormatter(logging.Formatter("[%(asctime)s] %(message)s"))
    logger.addHandler(_ch)
    logger.setLevel(logging.INFO)


# Page configuration  (must be the very first Streamlit call)
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="HirangAI",
    page_icon="H",
    layout="wide",
    initial_sidebar_state="expanded",
)






# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS — dark glassmorphism theme
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
<style>
/* Google Font */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* Global */
html, body, [class*="css"] { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; }
.stApp { background: #09090d; min-height: 100vh; }
#MainMenu, footer, header { display: none; }
.main .block-container { padding-top: 24px !important; }

/* Guarantee main-content expanders are fully interactive */
.main [data-testid="stExpander"] summary,
.main details > summary {
    cursor: pointer !important;
    pointer-events: auto !important;
    position: relative;
    z-index: 2;
}

/* Sticky nav */
.hirang-nav {
position: fixed; top: 0; left: 0; right: 0; z-index: 9999;
height: 48px; background: rgba(9,9,13,0.97);
border-bottom: 1px solid rgba(255,255,255,0.07);
padding: 0 24px; display: flex; align-items: center;
justify-content: space-between;
}
.hirang-nav-brand { font-size: 0.95rem; font-weight: 700; color: #f1f5f9; letter-spacing: -0.3px; }
.hirang-nav-links { display: flex; gap: 4px; }
.hirang-nav-links a {
color: #64748b; text-decoration: none; font-size: 0.8rem; font-weight: 500;
padding: 5px 10px; border-radius: 5px; transition: color 0.12s, background 0.12s;
}
.hirang-nav-links a:hover { color: #f1f5f9; background: rgba(255,255,255,0.05); }

/* Section anchor helper */
.section-anchor { display: block; height: 56px; margin-top: -56px; visibility: hidden; pointer-events: none; }



/* Section heading */
.section-heading {
font-size: 0.68rem; font-weight: 600; letter-spacing: 1.5px;
text-transform: uppercase; color: #4f46e5;
margin: 28px 0 12px; padding-bottom: 8px;
border-bottom: 1px solid rgba(79,70,229,0.15);
}

/* Cards */
.glass-card, .card {
background: #111118; border: 1px solid rgba(255,255,255,0.07);
border-radius: 8px; padding: 20px; margin-bottom: 12px;
transition: border-color 0.15s;
}
.glass-card:hover, .card:hover { border-color: rgba(79,70,229,0.2); }
.card-label { font-size: 0.67rem; font-weight: 600; letter-spacing: 1.2px; text-transform: uppercase; color: #4f46e5; margin-bottom: 4px; }
.card-title { font-size: 0.95rem; font-weight: 600; color: #f1f5f9; margin-bottom: 4px; }
.card-desc  { font-size: 0.8rem; color: #64748b; line-height: 1.5; margin-bottom: 14px; }

/* Info boxes */
.info-box    { background: rgba(79,70,229,0.05); border: 1px solid rgba(79,70,229,0.15); border-left: 2px solid #4f46e5; border-radius: 6px; padding: 10px 14px; color: #a5b4fc; font-size: 0.82rem; margin-top: 8px; line-height: 1.5; }
.success-box { background: rgba(5,150,105,0.05); border: 1px solid rgba(5,150,105,0.18); border-left: 2px solid #059669; border-radius: 6px; padding: 10px 14px; color: #6ee7b7; font-size: 0.82rem; margin-top: 8px; }
.warning-box { background: rgba(217,119,6,0.05); border: 1px solid rgba(217,119,6,0.18); border-left: 2px solid #d97706; border-radius: 6px; padding: 10px 14px; color: #fcd34d; font-size: 0.82rem; margin-top: 8px; }

/* Divider */
hr { border: none; border-top: 1px solid rgba(255,255,255,0.06) !important; margin: 20px 0 !important; }

/* Sidebar */
[data-testid="stSidebar"] { background: #0d0d14 !important; border-right: 1px solid rgba(255,255,255,0.06) !important; }
[data-testid="stSidebar"] .block-container { padding-top: 56px !important; }
.sidebar-logo    { font-size: 1.05rem; font-weight: 700; color: #f1f5f9; margin-bottom: 2px; }
.sidebar-tagline { font-size: 0.72rem; color: #475569; margin-bottom: 20px; }
.sidebar-stat { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.07); border-radius: 6px; padding: 10px 14px; margin-bottom: 8px; }
.sidebar-stat-label { font-size: 0.67rem; color: #64748b; font-weight: 500; letter-spacing: 0.5px; text-transform: uppercase; }
.sidebar-stat-value { font-size: 1.25rem; font-weight: 700; color: #f1f5f9; margin-top: 2px; }

/* Widgets */
.stTextArea > div > div > textarea {
background: #111118 !important; border: 1px solid rgba(255,255,255,0.1) !important;
border-radius: 6px !important; color: #f1f5f9 !important;
font-family: 'Inter', sans-serif !important; font-size: 0.875rem !important;
resize: vertical; transition: border-color 0.15s !important;
}
.stTextArea > div > div > textarea:focus {
border-color: rgba(79,70,229,0.5) !important;
box-shadow: 0 0 0 2px rgba(79,70,229,0.08) !important;
}
.stFileUploader { background: rgba(17,17,24,0.8) !important; border: 1px dashed rgba(255,255,255,0.1) !important; border-radius: 6px !important; }
.stFileUploader:hover { border-color: rgba(79,70,229,0.35) !important; }
.stButton > button {
background: #4f46e5 !important; border: none !important; border-radius: 6px !important;
color: #fff !important; font-family: 'Inter', sans-serif !important;
font-size: 0.875rem !important; font-weight: 600 !important; padding: 10px 20px !important;
transition: background 0.15s !important; box-shadow: none !important;
}
.stButton > button:hover { background: #4338ca !important; transform: none !important; box-shadow: none !important; }
label, .stFileUploader label { color: #64748b !important; font-size: 0.82rem !important; font-weight: 500 !important; }

/* Score bars */
@keyframes bar-grow { from { width: 0%; } to { width: var(--bar-w); } }
.cat-bar-track { background: rgba(255,255,255,0.05); border-radius: 3px; height: 6px; overflow: hidden; flex: 1; }
.cat-bar-fill  { height: 100%; border-radius: 3px; animation: bar-grow 0.5s ease-out forwards; width: var(--bar-w); }

/* ATS */
.ats-chip { display: inline-flex; align-items: center; padding: 2px 8px; border-radius: 4px; font-size: 0.7rem; font-weight: 500; margin: 2px 2px 2px 0; white-space: nowrap; }
.ats-chip-match { background: rgba(5,150,105,0.08); border: 1px solid rgba(5,150,105,0.22); color: #34d399; }
.ats-chip-miss  { background: rgba(217,119,6,0.07);  border: 1px solid rgba(217,119,6,0.2);  color: #fbbf24; }
.ats-section-row { display: flex; align-items: center; gap: 8px; padding: 5px 0; font-size: 0.8rem; border-bottom: 1px solid rgba(255,255,255,0.04); }
.ats-section-ok   { color: #34d399; }
.ats-section-miss { color: #fbbf24; }
.ats-sub-label { font-size: 0.72rem; font-weight: 600; color: #64748b; margin-bottom: 4px; }
.ats-bar-track { background: rgba(255,255,255,0.05); border-radius: 3px; height: 5px; overflow: hidden; flex: 1; }
.ats-bar-fill  { height: 100%; border-radius: 3px; animation: bar-grow 0.5s ease-out forwards; width: var(--bar-w); background: #0891b2; }

/* Status badges */
.status-badge { display: inline-flex; align-items: center; gap: 3px; padding: 2px 7px; border-radius: 4px; font-size: 0.65rem; font-weight: 600; white-space: nowrap; vertical-align: middle; }
.status-shortlisted { background: rgba(5,150,105,0.08); border: 1px solid rgba(5,150,105,0.25); color: #34d399; }
.status-hold        { background: rgba(217,119,6,0.08); border: 1px solid rgba(217,119,6,0.25); color: #fbbf24; }
.status-rejected    { background: rgba(220,38,38,0.08); border: 1px solid rgba(220,38,38,0.25); color: #f87171; }
.status-new         { background: rgba(100,116,139,0.06); border: 1px solid rgba(100,116,139,0.18); color: #64748b; }

/* Pipeline stat card */
.pipeline-stat-card { background: #111118; border: 1px solid rgba(255,255,255,0.07); border-radius: 8px; padding: 16px 18px; text-align: center; }
.pipeline-stat-num  { font-size: 1.75rem; font-weight: 700; line-height: 1; }
.pipeline-stat-lbl  { font-size: 0.65rem; font-weight: 600; letter-spacing: 1px; text-transform: uppercase; color: #64748b; margin-top: 4px; }

/* Interview rec */
.interview-rec-strip { display: flex; align-items: center; gap: 12px; border-radius: 6px; padding: 10px 14px; margin-bottom: 12px; }

/* Shortlist card */
.shortlist-cand-card { background: rgba(5,150,105,0.04); border: 1px solid rgba(5,150,105,0.12); border-left: 2px solid #059669; border-radius: 6px; padding: 10px 14px; margin-bottom: 6px; }

/* Weight badges */
.weight-total-ok  { display: flex; justify-content: space-between; align-items: center; background: rgba(5,150,105,0.05); border: 1px solid rgba(5,150,105,0.2); border-radius: 6px; padding: 8px 14px; margin: 10px 0 8px; }
.weight-total-err { display: flex; justify-content: space-between; align-items: center; background: rgba(220,38,38,0.05); border: 1px solid rgba(220,38,38,0.2); border-radius: 6px; padding: 8px 14px; margin: 10px 0 8px; }
.weight-total-label     { font-size: 0.7rem; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.5px; }
.weight-total-value-ok  { font-size: 0.88rem; font-weight: 700; color: #34d399; }
.weight-total-value-err { font-size: 0.88rem; font-weight: 700; color: #f87171; }

/* Getting Started workflow */
.workflow-card { background: #111118; border: 1px solid rgba(255,255,255,0.07); border-radius: 8px; padding: 16px; text-align: center; }
.workflow-step-num   { font-size: 0.65rem; font-weight: 700; color: #4f46e5; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 6px; }
.workflow-step-label { font-size: 0.82rem; font-weight: 500; color: #94a3b8; line-height: 1.4; }

/* Overview cards */
.top-candidate-card { background: #111118; border: 1px solid rgba(79,70,229,0.18); border-radius: 8px; padding: 22px; }
.overview-metric     { background: #111118; border: 1px solid rgba(255,255,255,0.06); border-radius: 8px; padding: 14px 16px; text-align: center; }
.overview-metric-num { font-size: 1.6rem; font-weight: 700; line-height: 1; }
.overview-metric-lbl { font-size: 0.65rem; font-weight: 600; letter-spacing: 1px; text-transform: uppercase; color: #64748b; margin-top: 4px; }

/* Disabled button state */
.stButton > button:disabled { opacity: 0.35 !important; cursor: not-allowed !important; }
.stButton > button:disabled:hover { background: #4f46e5 !important; transform: none !important; }

/* Pipeline status bar */
.pipeline-status { display: flex; align-items: center; gap: 0; padding: 10px 0 18px;
border-bottom: 1px solid rgba(255,255,255,0.06); margin-bottom: 20px; }

/* Compact upload list dot */
.upload-dot-ok   { width: 6px; height: 6px; border-radius: 50%; background: #059669; display: inline-block; }
.upload-dot-fail { width: 6px; height: 6px; border-radius: 50%; background: #dc2626; display: inline-block; }

/* Sidebar expander clean look */
[data-testid="stSidebar"] .stExpander { border: 1px solid rgba(255,255,255,0.07) !important;
border-radius: 6px !important; margin-top: 6px; }
[data-testid="stSidebar"] .stExpander summary { font-size: 0.8rem !important; color: #94a3b8 !important;
font-weight: 500 !important; }

/* Weight total badges */
.weight-total-ok, .weight-total-err { border-radius: 5px; padding: 7px 12px; margin: 8px 0; }

/* Tighter glass card */
.glass-card, .card { padding: 16px 18px !important; }

/* Job description card label + title spacing */
.card-label { margin-bottom: 2px !important; }
.card-title { margin-bottom: 6px !important; }
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Session state defaults
# ─────────────────────────────────────────────────────────────────────────────

if "analysis_count" not in st.session_state:
    st.session_state.analysis_count = 0
if "last_candidate" not in st.session_state:
    st.session_state.last_candidate = "—"
if "last_score" not in st.session_state:
    st.session_state.last_score = "—"
if "parsed_resume" not in st.session_state:
    st.session_state.parsed_resume = None
if "parsed_file_name" not in st.session_state:
    st.session_state.parsed_file_name = None
if "ranking_mode" not in st.session_state:
    st.session_state.ranking_mode = "Combined Score"
if "candidate_statuses" not in st.session_state:
    st.session_state.candidate_statuses = {}
if "pipeline_filter" not in st.session_state:
    st.session_state.pipeline_filter = "All"
if "trigger_analysis" not in st.session_state:
    st.session_state.trigger_analysis = False

# HR weight defaults (integers, must sum to 100)
if "hr_weights" not in st.session_state:
    st.session_state.hr_weights = {
        "technical_skills":          30,
        "relevant_experience":        30,
        "education":                  15,
        "certifications":             10,
        "communication_soft_skills":  15,
    }
for _k, _v in st.session_state.hr_weights.items():
    _skey = f"slider_{_k}"
    if _skey not in st.session_state:
        st.session_state[_skey] = _v

# ── Built-in evaluation templates ────────────────────────────────────────
BUILTIN_TEMPLATES: dict[str, dict] = {
    "Balanced Assessment": {
        "description": "Equal weight across all dimensions. Ideal for general hiring.",
        "weights": {
            "technical_skills":          30,
            "relevant_experience":        30,
            "education":                  15,
            "certifications":             10,
            "communication_soft_skills":  15,
        },
    },
    "Technical-Heavy Evaluation": {
        "description": "Prioritises technical depth. Best for engineering and developer roles.",
        "weights": {
            "technical_skills":          45,
            "relevant_experience":        25,
            "education":                  10,
            "certifications":             10,
            "communication_soft_skills":  10,
        },
    },
    "Experience-Focused Evaluation": {
        "description": "Emphasises work history and industry tenure over credentials.",
        "weights": {
            "technical_skills":          20,
            "relevant_experience":        45,
            "education":                  15,
            "certifications":             5,
            "communication_soft_skills":  15,
        },
    },
    "Skills-First Screening": {
        "description": "Combines technical and soft skills. Useful for fast-growing startups.",
        "weights": {
            "technical_skills":          35,
            "relevant_experience":        20,
            "education":                  10,
            "certifications":             10,
            "communication_soft_skills":  25,
        },
    },
    "Leadership & Communication Focus": {
        "description": "Prioritises people skills and leadership potential.",
        "weights": {
            "technical_skills":          15,
            "relevant_experience":        30,
            "education":                  15,
            "certifications":             5,
            "communication_soft_skills":  35,
        },
    },
    "Certification-Weighted Evaluation": {
        "description": "Favours formal qualifications. Suitable for regulated industries.",
        "weights": {
            "technical_skills":          20,
            "relevant_experience":        25,
            "education":                  20,
            "certifications":             25,
            "communication_soft_skills":  10,
        },
    },
}

_DEFAULT_TEMPLATE = "Balanced Assessment"

# Custom templates persisted across reruns
if "custom_templates" not in st.session_state:
    st.session_state.custom_templates = {}  # name -> {description, weights}
if "active_template" not in st.session_state:
    st.session_state.active_template = _DEFAULT_TEMPLATE
if "batch_results" not in st.session_state:
    st.session_state.batch_results = []  # list of {file, result}

def _reset_weights_cb():
    custom_tmpls = st.session_state.get("custom_templates", {})
    all_tmpls = {**BUILTIN_TEMPLATES, **custom_tmpls}
    active_tmpl = st.session_state.get("active_template", _DEFAULT_TEMPLATE)
    _reset_weights = dict(
        all_tmpls.get(active_tmpl, BUILTIN_TEMPLATES[_DEFAULT_TEMPLATE])["weights"]
    )
    st.session_state.hr_weights = _reset_weights
    for _k, _v in _reset_weights.items():
        st.session_state[f"slider_{_k}"] = _v

# ── Schema version guard ──────────────────────────────────────────────────────
# If a cached EvaluationResult is missing fields added in later schema
# versions, evict it so the user gets a fresh run instead of an AttributeError.
_stale = st.session_state.get("last_result")
if _stale is not None and (
    not hasattr(_stale, "category_scores")
    or not hasattr(_stale, "ats_result")
):
    del st.session_state["last_result"]

# =============================================================================
# RECRUITER EXPORT & COMPARISON HELPERS
# =============================================================================

from fpdf import FPDF

def safe_pdf_text(text):
    if not text:
        return ""
    if not isinstance(text, str):
        text = str(text)

    # Basic replacements of unsupported characters
    replacements = {
        "—": "-",
        "–": "-",
        "•": "*",
        "✓": "OK",
        "✔": "OK",
        "’": "'",
        "‘": "'",
        "”": '"',
        "“": '"',
        "…": "...",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    # Replaces tabs/newlines appropriately
    text = text.replace("\t", "    ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # Inserts break opportunities into extremely long tokens (longer than 40 chars)
    lines = text.split("\n")
    sanitized_lines = []
    for line in lines:
        words = line.split(" ")
        sanitized_words = []
        for word in words:
            if len(word) > 40:
                parts = [word[i:i+35] for i in range(0, len(word), 35)]
                sanitized_words.append(" ".join(parts))
            else:
                sanitized_words.append(word)
        sanitized_lines.append(" ".join(sanitized_words))
    text = "\n".join(sanitized_lines)

    # Limits excessively long strings
    if len(text) > 5000:
        text = text[:5000] + " ... [truncated]"

    # Encode as latin-1 and decode back safely
    return text.encode("latin-1", "replace").decode("latin-1")


def sanitize_pdf_text(text):
    return safe_pdf_text(text)


class SanitizedFPDF(FPDF):
    def cell(self, w, h=0, txt="", *args, **kwargs):
        if not txt:
            return False
        txt = sanitize_pdf_text(txt)
        if not txt:
            return False
        return super().cell(w, h, txt, *args, **kwargs)

    def multi_cell(self, w, h=0, txt="", *args, **kwargs):
        if not txt:
            return False
        txt = sanitize_pdf_text(txt)
        if not txt:
            return False
        return super().multi_cell(w, h, txt, *args, **kwargs)

    def write(self, h, txt="", *args, **kwargs):
        if not txt:
            return False
        txt = sanitize_pdf_text(txt)
        if not txt:
            return False
        return super().write(h, txt, *args, **kwargs)

    def text(self, x, y, txt="", *args, **kwargs):
        if not txt:
            return False
        txt = sanitize_pdf_text(txt)
        if not txt:
            return False
        return super().text(x, y, txt, *args, **kwargs)


def _score_colour_rgb(score: int | None) -> tuple[int, int, int]:
    """Return (R, G, B) matching the app's score colour scale."""
    if score is None:
        return (148, 163, 184)
    if score >= 85:
        return (16, 185, 129)
    if score >= 65:
        return (99, 102, 241)
    if score >= 40:
        return (245, 158, 11)
    return (239, 68, 68)


def _rec_colour_rgb(rec: str | None) -> tuple[int, int, int]:
    _map = {
        "Exceptional Candidate":      (16, 185, 129),
        "Outstanding Match":          (16, 185, 129),
        "Strong Hire":                (99, 102, 241),
        "Recommended":                (99, 102, 241),
        "Potential Fit":              (245, 158, 11),
        "Consider with Reservations": (245, 158, 11),
        "Weak Match":                 (239, 68, 68),
        "Not Recommended":            (239, 68, 68),
    }
    return _map.get(rec or "", (148, 163, 184))


def _build_rankings_pdf(
    ranked: list[dict],
    active_score_fn,
    template_name: str,
) -> bytes:
    """Generate a rankings PDF and return the raw bytes."""
    from prompts import CATEGORY_KEYS, CATEGORY_LABELS
    from candidate_evaluator import _score_to_recommendation as _s2r

    pdf = SanitizedFPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Title block ──────────────────────────────────────────────────────
    pdf.set_fill_color(13, 15, 26)
    pdf.rect(0, 0, 297, 40, "F")
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(129, 140, 248)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, "HirangAI — Candidate Rankings", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    pdf.set_x(10)
    pdf.cell(0, 6, f"Evaluation Template: {template_name}", ln=True)
    pdf.set_x(10)
    pdf.cell(0, 6, f"Total Candidates: {len(ranked)}", ln=True)

    pdf.set_y(46)

    # ── Table header ─────────────────────────────────────────────────────
    COLS = [
        ("Rank",          12),
        ("Candidate",     44),
        ("Score",         14),
        ("Recommendation",44),
        ("Confidence",    20),
        ("Tech",          14),
        ("Exp",           14),
        ("Edu",           14),
        ("Cert",          14),
        ("Soft",          14),
    ]
    pdf.set_fill_color(30, 32, 50)
    pdf.set_text_color(165, 180, 252)
    pdf.set_font("Helvetica", "B", 8)
    for label, w in COLS:
        pdf.cell(w, 8, label, border=0, fill=True, align="C")
    pdf.ln()

    # ── Table rows ───────────────────────────────────────────────────────
    medals = {1: "1st", 2: "2nd", 3: "3rd"}
    for rank, entry in enumerate(ranked, 1):
        r   = entry["result"]
        sc  = active_score_fn(entry)
        rec = _s2r(sc) if sc is not None else (r.recommendation or "—")
        conf = r.confidence_level or "Medium"
        name = r.candidate_name or entry["file"].replace(".pdf", "")
        cats = r.category_scores or {}

        # Row background alternates
        if rank % 2 == 0:
            pdf.set_fill_color(20, 22, 38)
        else:
            pdf.set_fill_color(16, 18, 30)

        sr, sg, sb = _score_colour_rgb(sc)
        rr, rg, rb = _rec_colour_rgb(rec)

        row_data = [
            (medals.get(rank, f"#{rank}"), 12, (148, 163, 184), "C"),
            (name[:28],                    44, (226, 232, 240), "L"),
            (str(sc) if sc is not None else "N/A", 14, (sr, sg, sb), "C"),
            (rec[:22],                     44, (rr, rg, rb), "L"),
            (conf,                         20, (148, 163, 184), "C"),
            (str(cats.get("technical_skills", "—")),          14, (148, 163, 184), "C"),
            (str(cats.get("relevant_experience", "—")),       14, (148, 163, 184), "C"),
            (str(cats.get("education", "—")),                 14, (148, 163, 184), "C"),
            (str(cats.get("certifications", "—")),            14, (148, 163, 184), "C"),
            (str(cats.get("communication_soft_skills", "—")), 14, (148, 163, 184), "C"),
        ]
        for text, w, (cr, cg, cb), align in row_data:
            pdf.set_text_color(cr, cg, cb)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(w, 7, text, border=0, fill=True, align=align)
        pdf.ln()

    # ── Footer ───────────────────────────────────────────────────────────
    pdf.set_y(-12)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 6, "Generated by HirangAI  •  Confidential — For Recruiter Use Only", align="C")

    return bytes(pdf.output())


def _build_rankings_excel(
    ranked: list[dict],
    active_score_fn,
    template_name: str,
) -> bytes:
    """Generate a rankings Excel workbook and return the raw bytes."""
    import io
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from candidate_evaluator import _score_to_recommendation as _s2r

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Rankings"

    # ── Palette ──────────────────────────────────────────────────────────
    BG_DARK   = "0D0F1A"
    BG_HEADER = "1E2032"
    BG_ROW_A  = "10121E"
    BG_ROW_B  = "141626"
    FG_PURPLE = "818CF8"
    FG_LIGHT  = "E2E8F0"
    FG_MUTED  = "64748B"

    thin = Side(style="thin", color="1E2032")
    border = Border(bottom=thin)

    def hex_fill(hex_col: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_col)

    def score_hex(score: int | None) -> str:
        if score is None: return "94A3B8"
        if score >= 85:   return "10B981"
        if score >= 65:   return "6366F1"
        if score >= 40:   return "F59E0B"
        return "EF4444"

    def rec_hex(rec: str | None) -> str:
        m = {
            "Exceptional Candidate": "10B981", "Outstanding Match": "10B981",
            "Strong Hire": "6366F1",           "Recommended": "6366F1",
            "Potential Fit": "F59E0B",         "Consider with Reservations": "F59E0B",
            "Weak Match": "EF4444",            "Not Recommended": "EF4444",
        }
        return m.get(rec or "", "94A3B8")

    # ── Title rows ───────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"] = "HirangAI — Candidate Rankings"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=FG_PURPLE)
    ws["A1"].fill = hex_fill(BG_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Template: {template_name}   |   Total Candidates: {len(ranked)}"
    ws["A2"].font = Font(name="Calibri", size=9, color=FG_MUTED)
    ws["A2"].fill = hex_fill(BG_DARK)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ───────────────────────────────────────────────────
    headers = [
        "Rank", "Candidate", "Email", "Contact Number", "File", "Overall Score", "Recommendation",
        "Confidence", "Technical", "Experience", "Education", "Certifications", "Soft Skills",
    ]
    ws.append([])  # row 3 blank
    ws.append(headers)  # row 4

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = Font(name="Calibri", size=9, bold=True, color=FG_PURPLE)
        cell.fill = hex_fill(BG_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[4].height = 20

    # ── Data rows ────────────────────────────────────────────────────────
    medals = {1: "🥇 1st", 2: "🥈 2nd", 3: "🥉 3rd"}
    for rank, entry in enumerate(ranked, 1):
        r    = entry["result"]
        sc   = active_score_fn(entry)
        rec  = _s2r(sc) if sc is not None else (r.recommendation or "—")
        conf = r.confidence_level or "Medium"
        name = r.candidate_name or entry["file"].replace(".pdf", "")
        cats = r.category_scores or {}

        # Heuristic extraction of contact info with parsed resume fallback
        email = getattr(r, "email", None)
        phone = getattr(r, "phone", None)
        if not email:
            parsed_res = st.session_state.get(f"parsed_{entry['file']}")
            if parsed_res:
                email = getattr(parsed_res, "email", None)
        if not phone:
            parsed_res = st.session_state.get(f"parsed_{entry['file']}")
            if parsed_res:
                phone = getattr(parsed_res, "phone", None)

        email = email or "Not Found"
        phone = phone or "Not Found"

        row_vals = [
            medals.get(rank, f"#{rank}"),
            name,
            email,
            phone,
            entry["file"],
            sc if sc is not None else "N/A",
            rec,
            conf,
            cats.get("technical_skills", ""),
            cats.get("relevant_experience", ""),
            cats.get("education", ""),
            cats.get("certifications", ""),
            cats.get("communication_soft_skills", ""),
        ]
        ws.append(row_vals)
        data_row = ws.max_row
        bg = BG_ROW_A if rank % 2 == 1 else BG_ROW_B
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = hex_fill(bg)
            cell.font = Font(name="Calibri", size=9, color=FG_LIGHT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Colour-code score and recommendation cells (shifted due to inserted columns)
        score_cell = ws.cell(row=data_row, column=6)
        score_cell.font = Font(name="Calibri", size=9, bold=True, color=score_hex(sc))

        rec_cell = ws.cell(row=data_row, column=7)
        rec_cell.font = Font(name="Calibri", size=9, bold=True, color=rec_hex(rec))

        ws.row_dimensions[data_row].height = 16

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = [10, 28, 28, 20, 28, 14, 28, 12, 12, 12, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze header row ────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_individual_pdf(
    entry: dict,
    active_score: int | None,
    active_recommendation: str,
    hr_weights: dict,
) -> bytes:
    """Generate a detailed individual candidate report PDF."""
    from prompts import CATEGORY_KEYS, CATEGORY_LABELS
    import textwrap

    result = entry["result"]
    name   = result.candidate_name or entry["file"].replace(".pdf", "")

    pdf = SanitizedFPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    # ── Header banner ────────────────────────────────────────────────────
    pdf.set_fill_color(13, 15, 26)
    pdf.rect(0, 0, 210, 38, "F")
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(129, 140, 248)
    pdf.set_xy(12, 6)
    pdf.cell(0, 10, "HirangAI", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 116, 139)
    pdf.set_x(12)
    pdf.cell(0, 7, "Candidate Evaluation Report  —  Confidential", ln=True)
    pdf.set_x(12)
    pdf.cell(0, 7, f"For Recruiter Use Only", ln=True)

    pdf.set_y(44)

    # ── Candidate identity block ─────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(226, 232, 240)
    pdf.cell(0, 9, name, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    contact_parts = []
    if result.email: contact_parts.append(result.email)
    if result.phone: contact_parts.append(result.phone)
    if contact_parts:
        pdf.cell(0, 6, "  |  ".join(contact_parts), ln=True)
    pdf.ln(4)

    # ── Score / Recommendation / Confidence row ──────────────────────────
    sr, sg, sb = _score_colour_rgb(active_score)
    rr, rg, rb = _rec_colour_rgb(active_recommendation)

    # Score box
    pdf.set_fill_color(sr // 6, sg // 6, sb // 6)
    pdf.set_text_color(sr, sg, sb)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(40, 7, "OVERALL SCORE", border=0, fill=True, align="C")
    pdf.set_text_color(rr, rg, rb)
    pdf.cell(80, 7, "RECOMMENDATION", border=0, fill=False, align="C")
    pdf.set_text_color(148, 163, 184)
    pdf.cell(40, 7, "CONFIDENCE", border=0, fill=False, align="C")
    pdf.ln()
    pdf.set_text_color(sr, sg, sb)
    pdf.set_font("Helvetica", "B", 22)
    sc_str = f"{active_score}/100" if active_score is not None else "N/A"
    pdf.cell(40, 12, sc_str, align="C")
    pdf.set_text_color(rr, rg, rb)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(80, 12, active_recommendation, align="C")
    pdf.set_text_color(148, 163, 184)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(40, 12, result.confidence_level or "Medium", align="C")
    pdf.ln(16)

    def section_heading(title: str) -> None:
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(99, 102, 241)
        pdf.cell(0, 6, title.upper(), ln=True)
        pdf.set_draw_color(99, 102, 241)
        pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 186, pdf.get_y())
        pdf.ln(3)

    def body_text(text: str, colour=(148, 163, 184)) -> None:
        if not text:
            return
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*colour)
        pdf.set_x(pdf.l_margin)
        w_safe = pdf.w - pdf.l_margin - pdf.r_margin
        pdf.multi_cell(w_safe, 5, text)
        pdf.ln(1)

    def bullet_list(items: list[str], colour=(148, 163, 184)) -> None:
        if not items:
            return
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*colour)
        for item in items:
            if not item:
                continue
            pdf.set_x(pdf.l_margin)
            wrapped = textwrap.fill(f"  *  {item}", width=95)
            w_safe = pdf.w - pdf.l_margin - pdf.r_margin
            pdf.multi_cell(w_safe, 5, wrapped)
        pdf.ln(2)

    # ── Category Breakdown ───────────────────────────────────────────────
    if result.category_scores:
        section_heading("Category Breakdown")
        cat_icons = {
            "technical_skills": "Tech Skills",
            "relevant_experience": "Relevant Exp.",
            "education": "Education",
            "certifications": "Certifications",
            "communication_soft_skills": "Soft Skills",
        }
        for key in CATEGORY_KEYS:
            label  = cat_icons.get(key, CATEGORY_LABELS.get(key, key))
            s      = result.category_scores.get(key)
            w_pct  = hr_weights.get(key, 0)
            s_str  = str(s) if s is not None else "—"
            sr2, sg2, sb2 = _score_colour_rgb(s)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(203, 213, 225)
            pdf.cell(60, 6, label)
            pdf.set_text_color(sr2, sg2, sb2)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(20, 6, s_str, align="R")
            pdf.set_text_color(100, 116, 139)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(0, 6, f"  (weight: {w_pct}%)", ln=True)
        pdf.ln(3)

    # ── Strengths ────────────────────────────────────────────────────────
    if result.strengths:
        section_heading("Key Strengths")
        bullet_list(result.strengths, colour=(110, 231, 183))

    # ── Skill Gaps ───────────────────────────────────────────────────────
    if result.skill_gaps:
        section_heading("Skill Gaps")
        bullet_list(result.skill_gaps, colour=(248, 113, 113))

    # ── Experience Summary ───────────────────────────────────────────────
    if result.experience_summary:
        section_heading("Experience Summary")
        body_text(result.experience_summary)

    # ── Full AI Analysis ─────────────────────────────────────────────────
    if result.raw_llm_response:
        section_heading("Full AI Analysis")
        # Strip markdown bold/heading syntax for plain PDF text
        import re
        clean = re.sub(r"\*{1,3}(.+?)\*{1,3}", r"\1", result.raw_llm_response)
        clean = re.sub(r"#{1,4}\s*", "", clean)
        clean = re.sub(r"```[a-z]*", "", clean)
        clean = clean.replace("`", "")
        # Limit to ~3000 chars to avoid enormous PDFs
        if len(clean) > 3000:
            clean = clean[:3000] + "\n\n[… truncated — see full analysis in the app …]"
        body_text(clean.strip())

    # ── Footer ───────────────────────────────────────────────────────────
    pdf.set_y(-12)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 6, "Generated by HirangAI  •  Confidential — For Recruiter Use Only", align="C")

    return bytes(pdf.output())


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────


# Sticky navigation header commented out to see if it fixes the click blockage
# st.markdown(
#     """
# <div class="hirang-nav">
# <span class="hirang-nav-brand">HirangAI</span>
# <div class="hirang-nav-links">
# <a href="#overview">Overview</a>
# <a href="#leaderboard">Leaderboard</a>
# <a href="#evaluations">Evaluations</a>
# <a href="#comparison">Comparison</a>
# <a href="#export">Export</a>
# </div>
# </div>
# """,
#     unsafe_allow_html=True,
# )

with st.sidebar:
    st.markdown('<div class="sidebar-logo">HirangAI</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-tagline">Choose Better. Hire Smarter.</div>', unsafe_allow_html=True)

    st.markdown(
        f'''<div style="font-size:0.75rem;color:#64748b;margin-bottom:20px;padding-left:2px;">'''
        f'''Session — <strong>{st.session_state.analysis_count}</strong> candidates screened</div>''',
        unsafe_allow_html=True
    )

    # ── Assessment Profile ─────────────────────────────────────────
    st.markdown('<div class="section-heading">Assessment Profile</div>', unsafe_allow_html=True)

    _all_templates: dict[str, dict] = {**BUILTIN_TEMPLATES, **st.session_state.custom_templates}
    _template_names = list(_all_templates.keys())

    _current_idx = (
        _template_names.index(st.session_state.active_template)
        if st.session_state.active_template in _template_names
        else 0
    )

    _selected = st.selectbox(
        label="Assessment Profile",
        options=_template_names,
        index=_current_idx,
        key="template_selectbox",
        help="Weights recalculate scores automatically.",
        label_visibility="collapsed",
    )

    # Apply template when selection changes
    if _selected != st.session_state.active_template:
        new_weights = dict(_all_templates[_selected]["weights"])
        st.session_state.active_template = _selected
        st.session_state.hr_weights      = new_weights
        for _k, _v in new_weights.items():
            st.session_state[f"slider_{_k}"] = _v
        st.rerun()

    _tmpl_desc = _all_templates[_selected].get("description", "")
    if _tmpl_desc:
        st.markdown(
            f'''<div style="color:#64748b;font-size:0.75rem;margin:4px 0 12px;line-height:1.5;">{_tmpl_desc}</div>''',
            unsafe_allow_html=True,
        )

    # ── Advanced Configuration (collapsed by default) ──────────────
    # Show expanded if a custom template is active or it was manually opened
    _is_custom = _selected in st.session_state.custom_templates
    _adv_expanded = _is_custom or st.session_state.get("_adv_config_open", False)

    with st.expander("Advanced Configuration", expanded=_adv_expanded):
        st.markdown(
            '<div style="color:#64748b;font-size:0.72rem;margin-bottom:10px;">'
            'Adjust category weights. Total must equal 100%.</div>',
            unsafe_allow_html=True,
        )

        _cat_config = [
            ("technical_skills",         "Technical Skills"),
            ("relevant_experience",       "Relevant Experience"),
            ("education",                 "Education"),
            ("certifications",            "Certifications"),
            ("communication_soft_skills", "Soft Skills"),
        ]

        _w = st.session_state.hr_weights
        for _key, _label in _cat_config:
            _w[_key] = st.slider(
                label=_label,
                min_value=0,
                max_value=100,
                step=5,
                key=f"slider_{_key}",
            )

        _total = sum(_w.values())
        if _total == 100:
            st.markdown(
                '<div class="weight-total-ok">'
                '<span class="weight-total-label">Total Weight</span>'
                '<span class="weight-total-value-ok">100%</span>'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'''<div class="weight-total-err">
<span class="weight-total-label">Total Weight</span>
<span class="weight-total-value-err">{_total}%</span>
</div>''',
                unsafe_allow_html=True,
            )
            st.markdown(
                '<div style="color:#f87171;font-size:0.72rem;margin-top:4px;">'
                'Weights must sum to 100%.</div>',
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)
        _btn_reset, _btn_save = st.columns(2, gap="small")
        with _btn_reset:
            st.button(
                "Reset",
                key="reset_weights",
                use_container_width=True,
                on_click=_reset_weights_cb,
            )
        with _btn_save:
            _save_clicked = st.button("Save Template", key="open_save_template", use_container_width=True)

        if _save_clicked or st.session_state.get("_save_template_open"):
            st.session_state["_save_template_open"] = True
            _new_name = st.text_input(
                "Template name",
                placeholder="e.g. Senior Dev Profile",
                key="new_template_name",
            )
            _new_desc = st.text_input(
                "Description (optional)",
                placeholder="What does this template prioritise?",
                key="new_template_desc",
            )
            _do_save, _do_cancel = st.columns(2, gap="small")
            with _do_save:
                if st.button("Save", key="confirm_save_template", use_container_width=True):
                    if _new_name.strip():
                        st.session_state.custom_templates[_new_name.strip()] = {
                            "description": _new_desc.strip() or "Custom template",
                            "weights":     dict(_w),
                        }
                        st.session_state.active_template = _new_name.strip()
                        st.session_state["_save_template_open"] = False
                        st.rerun()
                    else:
                        st.markdown(
                            '<div style="color:#f87171;font-size:0.74rem;">'
                            'Enter a name for the template.</div>',
                            unsafe_allow_html=True,
                        )
            with _do_cancel:
                if st.button("Cancel", key="cancel_save_template", use_container_width=True):
                    st.session_state["_save_template_open"] = False
                    st.rerun()

        if st.session_state.custom_templates:
            st.markdown("<br>", unsafe_allow_html=True)
            _to_delete = st.selectbox(
                "Delete template",
                options=["— select to delete —"] + list(st.session_state.custom_templates.keys()),
                key="delete_template_select",
            )
            if _to_delete != "— select to delete —":
                if st.button(f"Delete \u2018{_to_delete}\u2019", key="confirm_delete_template"):
                    del st.session_state.custom_templates[_to_delete]
                    if st.session_state.active_template == _to_delete:
                        st.session_state.active_template = _DEFAULT_TEMPLATE
                        st.session_state.hr_weights = dict(
                            BUILTIN_TEMPLATES[_DEFAULT_TEMPLATE]["weights"]
                        )
                    st.rerun()

    # ── AI Status ─────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:0.72rem;color:#64748b;padding:8px 0;border-top:1px solid rgba(255,255,255,0.05);">'
        '<span style="color:#059669;font-weight:600;">Active</span> — Groq llama-3.3-70b</div>',
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# ── Pipeline status bar ──────────────────────────────────────────────────────
_jd_ready     = bool(st.session_state.get("job_description_input", "").strip())
_files_ready  = bool(st.session_state.get("resume_uploader"))
_profile_name = st.session_state.get("active_template", "Balanced Assessment")
_results_ready = bool(st.session_state.get("batch_results"))

def _step_style(active: bool) -> str:
    return (
        "font-size:0.78rem;font-weight:600;color:#f1f5f9;"
        if active else
        "font-size:0.78rem;font-weight:400;color:#475569;"
    )

_steps_html = (
    f'''<div style="display:flex;align-items:center;gap:0;padding:10px 0 18px;border-bottom:1px solid rgba(255,255,255,0.06);margin-bottom:20px;">'''
    f'''<span style="{_step_style(_jd_ready)}">Job Description</span>'''
    f'''<span style="margin:0 10px;color:#334155;font-size:0.75rem;">·</span>'''
    f'''<span style="{_step_style(_files_ready)}">Candidates</span>'''
    f'''<span style="margin:0 10px;color:#334155;font-size:0.75rem;">·</span>'''
    f'''<span style="{_step_style(True)}">Assessment Profile</span>'''
    f'''<span style="margin:0 10px;color:#334155;font-size:0.75rem;">·</span>'''
    f'''<span style="{_step_style(_jd_ready and _files_ready)}">Analyze</span>'''
    f'''<span style="margin:0 10px;color:#334155;font-size:0.75rem;">·</span>'''
    f'''<span style="{_step_style(_results_ready)}">Review</span>'''
    f'''</div>'''
)
st.markdown(_steps_html, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Main input layout
# ─────────────────────────────────────────────────────────────────────────────

left_col, right_col = st.columns([1.1, 0.9], gap="large")

# ── LEFT: Job Description ────────────────────────────────────────────────────
with left_col:
    st.markdown(
        """
<div class="glass-card">
<div class="card-label">Step 1</div>
<div class="card-title">Job Description</div>
<div class="card-desc">
Paste the full job posting — including responsibilities,
required skills, and qualifications.
</div>
</div>
""",
        unsafe_allow_html=True,
    )

    job_description = st.text_area(
        label="Job Description",
        placeholder=(
            "e.g.\n\n"
            "We are looking for a Senior Python Developer with 5+ years of experience...\n\n"
            "Responsibilities:\n"
            "• Design and implement scalable backend services\n"
            "• Collaborate with cross-functional teams\n\n"
            "Requirements:\n"
            "• Proficiency in Python, FastAPI, PostgreSQL\n"
            "• Experience with cloud platforms (AWS / GCP)\n"
            "• Strong communication skills"
        ),
        height=320,
        label_visibility="collapsed",
        key="job_description_input",
    )

    jd_word_count = len(job_description.split()) if job_description.strip() else 0
    if job_description.strip():
        st.markdown(
            f'<div style="font-size:0.78rem;color:#64748b;margin-top:6px;">'
            f'<strong style="color:#f1f5f9;">{jd_word_count}</strong> words · Job description ready</div>',
            unsafe_allow_html=True,
        )

# ── RIGHT: Resume Upload ─────────────────────────────────────────────────────
with right_col:
    st.markdown(
        """
<div class="glass-card">
<div class="card-label">Step 2</div>
<div class="card-title">Upload Resumes</div>
<div class="card-desc">
Upload one or more PDF resumes.
All candidates will be evaluated against the same job description.
</div>
</div>
""",
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader(
        label="Upload PDF Resumes",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload one or more PDF files. Each resume will be evaluated separately.",
        key="resume_uploader",
        label_visibility="collapsed",
    )

    # Normalise: file_uploader with accept_multiple_files returns a list
    if uploaded_files is None:
        uploaded_files = []

    if uploaded_files:
        # Parse any file not yet cached
        for _uf in uploaded_files:
            if st.session_state.get(f"parsed_{_uf.name}") is None:
                with st.spinner(f"Parsing {_uf.name}..."):
                    st.session_state[f"parsed_{_uf.name}"] = parse_resume(_uf.read())

        _n_ok = sum(
            1 for f in uploaded_files
            if st.session_state.get(f"parsed_{f.name}") and st.session_state[f"parsed_{f.name}"].is_valid
        )
        _n_fail = len(uploaded_files) - _n_ok

        # Clean candidate list header
        st.markdown(
            f'''<div style="margin-top:12px;">'''
f'''<div style="font-size:0.72rem;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:#64748b;margin-bottom:8px;">Uploaded Candidates</div>'''
f'''<div style="font-size:0.82rem;color:#94a3b8;margin-bottom:10px;">'''
f'''<span style="color:#f1f5f9;font-weight:600;">{len(uploaded_files)}</span> uploaded'''
f'''&nbsp;·&nbsp;'''
f'''<span style="color:#059669;font-weight:600;">{_n_ok}</span> ready'''
f'''{f"&nbsp;·&nbsp;<span style='color:#dc2626;font-weight:600;'>{_n_fail}</span> failed" if _n_fail else ""}'''
f'''</div>'''
f'''</div>''',
            unsafe_allow_html=True,
        )

        # Compact filename list
        for _uf in uploaded_files:
            _p = st.session_state.get(f"parsed_{_uf.name}")
            _ok = _p and _p.is_valid
            _dot_color = "#059669" if _ok else "#dc2626"
            _name_color = "#e2e8f0" if _ok else "#94a3b8"
            _err_msg = f'<span style="font-size:0.7rem;color:#dc2626;margin-left:auto;">Could not parse</span>' if not _ok else ""
            st.markdown(
                f'''<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.04);">'''
f'''<span style="width:6px;height:6px;border-radius:50%;background:{_dot_color};display:inline-block;flex-shrink:0;"></span>'''
f'''<span style="font-size:0.82rem;color:{_name_color};font-weight:500;">{_uf.name}</span>'''
f'''{_err_msg}'''
f'''</div>''',
                unsafe_allow_html=True,
            )

        # Collapsible details
        with st.expander("Show file details", expanded=False):
            for _uf in uploaded_files:
                _p  = st.session_state.get(f"parsed_{_uf.name}")
                _kb = round(_uf.size / 1024, 1)
                if _p and _p.is_valid:
                    st.markdown(
                        f'''<div style="font-size:0.78rem;color:#94a3b8;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.04);">'''
f'''<strong style="color:#e2e8f0;">{_uf.name}</strong>'''
f'''&nbsp;·&nbsp;{_p.page_count} page(s)'''
f'''&nbsp;·&nbsp;{_p.word_count} words'''
f'''&nbsp;·&nbsp;{_kb} KB'''
f'''</div>''',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f'''<div style="font-size:0.78rem;color:#dc2626;padding:4px 0;">'''
f'''{_uf.name} — could not extract text'''
f'''</div>''',
                        unsafe_allow_html=True,
                    )
    else:
        # Clear cached parses when all files removed
        for _k in [k for k in st.session_state if k.startswith("parsed_") and k not in ("parsed_resume", "parsed_file_name")]:
            del st.session_state[_k]
        st.markdown(
            '''<div style="font-size:0.82rem;color:#475569;margin-top:10px;">'''
'''No resumes uploaded yet.'''
'''</div>''',
            unsafe_allow_html=True,
        )



# ─────────────────────────────────────────────────────────────────────────────
# Analyse button + evaluation pipeline
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

# Determine readiness state
_uploaded_files_safe = uploaded_files if uploaded_files else []

# ── FIX: derive _n_valid from PERSISTENT parsed_* session state keys, not the ──
# ── live file_uploader return value.  The uploader returns [] on reruns that   ──
# ── were not triggered by a file-submit event (e.g. after st.rerun()), so      ──
# ── using it as the source makes _files_ok transiently False and silently      ──
# ── swallows the pipeline on those cycles.  Session state always persists.     ──
_SKIP_PARSED_KEYS = {"parsed_resume", "parsed_file_name"}
_n_valid = sum(
    1 for k, v in st.session_state.items()
    if k.startswith("parsed_")
    and k not in _SKIP_PARSED_KEYS
    and v is not None
    and v.is_valid
)

_jd_ok    = bool(job_description.strip())
_files_ok = _n_valid > 0

# Strict assessment weights sum validation
_hr_w = st.session_state.get("hr_weights", {})
_weights_sum = sum(_hr_w.values())
_weights_ok = (_weights_sum == 100)

_can_run = _jd_ok and _files_ok and _weights_ok

# Validation message shown above button
if not _jd_ok:
    st.markdown(
        '<div style="font-size:0.8rem;color:#64748b;margin-bottom:8px;padding:8px 12px;'
        'background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:6px;">'
        'Enter a job description to continue.</div>',
        unsafe_allow_html=True,
    )
elif not _files_ok:
    st.markdown(
        '<div style="font-size:0.8rem;color:#64748b;margin-bottom:8px;padding:8px 12px;'
        'background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:6px;">'
        'Upload at least one PDF resume to continue.</div>',
        unsafe_allow_html=True,
    )
elif not _weights_ok:
    st.markdown(
        f'<div style="font-size:0.8rem;color:#f87171;margin-bottom:8px;padding:8px 12px;'
        f'background:rgba(220,38,38,0.05);border:1px solid rgba(220,38,38,0.15);border-radius:6px;">'
        f'⚠️ Analysis is blocked because the total assessment category weights must sum to exactly 100% '
        f'(currently {_weights_sum}%). Please adjust the weights in the Advanced Configuration sidebar expander.</div>',
        unsafe_allow_html=True,
    )

_btn_label = (
    f"Analyze {_n_valid} Candidate{'s' if _n_valid != 1 else ''}"
    if _n_valid > 0 else "Analyze Candidates"
)


# Active production Analyze Candidates button (utilizing the verified working key)
analyse_clicked = st.button(
    _btn_label,
    key="analyze_candidates_test_btn",
    disabled=False,
    use_container_width=True,
    type="primary",
)

# Original button implementation hidden under if False for rollback purposes
if False:
    _btn_label_old = (
        f"Analyze {_n_valid} Candidate{'s' if _n_valid != 1 else ''}"
        if _n_valid > 0 else "Analyze Candidates"
    )
    analyse_clicked_old = st.button(
        _btn_label_old,
        key="analyze_pipeline_test_btn_old_hidden",  # Renamed key inside dead block to prevent Streamlit widget key collision!
        disabled=False,
        use_container_width=True,
        type="primary",
    )

pipeline_triggered = analyse_clicked

if pipeline_triggered:
    if _can_run:
        _valid_pairs = [
            (f, st.session_state[f"parsed_{f.name}"])
            for f in _uploaded_files_safe
            if st.session_state.get(f"parsed_{f.name}") and st.session_state[f"parsed_{f.name}"].is_valid
        ]

        if not _valid_pairs:
            st.error("No resumes could be parsed. Please upload valid PDF files.")
        else:
            _batch_results: list[dict] = []
            _progress = st.progress(0, text="Starting evaluation...")
            _status_placeholder = st.empty()

            for _i, (_uf, _parsed) in enumerate(_valid_pairs):
                _pct  = _i / len(_valid_pairs)
                _name = _parsed.candidate_name or _uf.name.replace(".pdf", "")
                _progress.progress(_pct, text=f"Evaluating {_name} ({_i + 1} of {len(_valid_pairs)})...")
                
                _status_placeholder.markdown(
                    f'''<div style="font-size:0.8rem;color:#64748b;padding:6px 0;">Analyzing <strong style="color:#f1f5f9;">{_name}</strong>...</div>''',
                    unsafe_allow_html=True,
                )
                
                try:
                    import time
                    _t_start = time.time()
                    _res = evaluate_candidate(
                        job_description=job_description,
                        parsed_resume=_parsed,
                    )
                    _t_elapsed = time.time() - _t_start
                    logger.info(
                        f"[HirangAI DEBUG] Groq API call succeeded for {_name} in {_t_elapsed:.2f}s."
                    )
                except Exception as _eval_err:
                    _t_elapsed = time.time() - _t_start
                    logger.error(
                        f"[HirangAI DEBUG] Groq API call failed for {_name}: {_eval_err}",
                        exc_info=True
                    )
                    from candidate_evaluator import EvaluationResult
                    _res = EvaluationResult(
                        candidate_name=_name,
                        match_score=None,
                        recommendation=None,
                        confidence_level=None,
                        strengths=[],
                        skill_gaps=[],
                        experience_summary="",
                        raw_llm_response="",
                        errors=[f"Evaluation failed: {_eval_err}"],
                    )
                
                _batch_results.append({"file": _uf.name, "result": _res})

            _progress.progress(1.0, text=f"Complete — {len(_batch_results)} candidate(s) analyzed")
            _status_placeholder.empty()

            # Result Storage
            logger.info("[HirangAI DEBUG] Storing batch results in session state.")
            
            st.session_state.batch_results  = _batch_results
            st.session_state.last_result    = _batch_results[-1]["result"]
            st.session_state.last_file_name = _batch_results[-1]["file"]

            _successful = [r for r in _batch_results if not r["result"].errors and r["result"].recommendation]
            st.session_state.analysis_count += len(_successful)
            if _successful:
                _top = max(_successful, key=lambda r: r["result"].match_score or 0)
                st.session_state.last_candidate = _top["result"].candidate_name or _top["file"]
                _ts  = _top["result"].match_score
                st.session_state.last_score = f"{_ts}/100" if _ts is not None else _top["result"].recommendation
            
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Result display  (reads from session_state — persists across reruns)
# ─────────────────────────────────────────────────────────────────────────────

if st.session_state.get("batch_results"):
    from prompts import CATEGORY_KEYS

    _br = st.session_state.batch_results
    _hr_w     = st.session_state.hr_weights
    _hr_total = sum(_hr_w.values())

    # Compute active (HR-adjusted) HirangAI score for every result
    def _active_score(entry: dict) -> int | None:
        r = entry["result"]
        if hasattr(r, "category_scores") and r.category_scores and _hr_total == 100:
            ws = sum(r.category_scores.get(k, 0) * (_hr_w.get(k, 0) / 100) for k in CATEGORY_KEYS)
            return max(0, min(100, round(ws)))
        return r.match_score

    # Combined Score = Fit × 0.70 + ATS × 0.30
    _TL_WEIGHT  = 0.70
    _ATS_WEIGHT = 0.30

    def _combined_score(entry: dict) -> int | None:
        tl  = _active_score(entry)
        _ar = getattr(entry["result"], "ats_result", None)
        ats = _ar.ats_score if _ar else None
        if tl is None and ats is None:
            return None
        t = tl  if tl  is not None else 0
        a = ats if ats is not None else 0
        if tl is None:  return a
        if ats is None: return t
        return max(0, min(100, round(t * _TL_WEIGHT + a * _ATS_WEIGHT)))

    def _ranking_score(entry: dict) -> int | None:
        mode = st.session_state.get("ranking_mode", "Combined Score")
        if mode == "Candidate Fit Score Only":
            return _active_score(entry)
        elif mode == "ATS Score Only":
            _ar = getattr(entry["result"], "ats_result", None)
            return _ar.ats_score if _ar else None
        else:  # Combined Score
            return _combined_score(entry)

    # Sort by ranking score descending, errors last
    _ranked = sorted(
        _br,
        key=lambda e: (_ranking_score(e) or 0) if not e["result"].errors else -1,
        reverse=True,
    )

    # ── Pipeline status helpers ──────────────────────────────────────────────
    _statuses = st.session_state.candidate_statuses

    def _get_status(entry):
        return _statuses.get(entry["file"], None)

    def _interview_rec(final_score):
        if final_score is None:   return "—", "#64748b"
        if final_score >= 80:     return "Strongly Recommended", "#10b981"
        elif final_score >= 65:   return "Recommended", "#6366f1"
        elif final_score >= 50:   return "Conditional", "#f59e0b"
        else:                     return "Not Recommended", "#ef4444"

    _n_shortlisted = sum(1 for e in _ranked if _get_status(e) == "shortlisted")
    _n_hold        = sum(1 for e in _ranked if _get_status(e) == "hold")
    _n_rejected    = sum(1 for e in _ranked if _get_status(e) == "rejected")
    _n_total       = len(_ranked)

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<span class="section-anchor" id="overview"></span>', unsafe_allow_html=True)

    # ── Pipeline Overview ────────────────────────────────────────────────────
    st.markdown('<div class="section-heading">Executive Summary</div>', unsafe_allow_html=True)
    
    # Identify top candidate
    if _ranked:
        _top_entry = _ranked[0]
        _top_res = _top_entry["result"]
        _top_name = _top_res.candidate_name or _top_entry["file"].replace(".pdf", "")
        _top_score = _ranking_score(_top_entry)
        _top_rec = _interview_rec(_top_score)[0] if _top_score is not None else (_top_res.recommendation or "Recommended")
        _top_conf = _top_res.confidence_level or "High"
        
        # Get 3-5 strengths
        _top_strengths = _top_res.strengths[:4] if _top_res.strengths else ["Demonstrates solid matching qualifications for the role."]
        _strengths_li = "".join(f'<li style="margin-bottom:4px;">{s}</li>' for s in _top_strengths)
        
        _avg_score = round(sum((_ranking_score(e) or 0) for e in _ranked if not e["result"].errors) / max(1, sum(1 for e in _ranked if not e["result"].errors)))
        _highest_score = _top_score if _top_score is not None else 0
        
        _exec_col1, _exec_col2 = st.columns([0.62, 0.38], gap="medium")
        with _exec_col1:
            st.markdown(
                f'''
<div class="top-candidate-card" style="height: 100%;">
<div class="card-label">Top Ranked Candidate</div>
<div style="font-size: 1.25rem; font-weight: 700; color: #f1f5f9; margin-bottom: 8px;">{_top_name}</div>
<div style="display: flex; gap: 16px; margin-bottom: 14px;">
<div>
<span style="font-size: 0.68rem; color: #64748b; text-transform: uppercase; font-weight: 600;">Match Score</span>
<div style="font-size: 1.5rem; font-weight: 800; color: #4f46e5;">{_top_score if _top_score is not None else 'N/A'}<span style="font-size: 0.85rem; font-weight: 500; color: #64748b;">/100</span></div>
</div>
<div>
<span style="font-size: 0.68rem; color: #64748b; text-transform: uppercase; font-weight: 600;">Recommendation</span>
<div style="font-size: 0.95rem; font-weight: 700; color: #059669; margin-top: 4px;">{_top_rec}</div>
</div>
<div>
<span style="font-size: 0.68rem; color: #64748b; text-transform: uppercase; font-weight: 600;">AI Confidence</span>
<div style="font-size: 0.95rem; font-weight: 700; color: #f1f5f9; margin-top: 4px;">{_top_conf}</div>
</div>
</div>
<div style="font-size: 0.72rem; color: #64748b; font-weight: 600; text-transform: uppercase; margin-bottom: 6px; letter-spacing: 0.5px;">Why This Candidate Ranked First:</div>
<ul style="color: #94a3b8; font-size: 0.8rem; line-height: 1.5; padding-left: 16px; margin: 0;">{_strengths_li}</ul>
</div>
''',
                unsafe_allow_html=True
            )
        with _exec_col2:
            # Summary Metrics Row
            st.markdown(
                f'''
<div style="display: flex; flex-direction: column; gap: 12px; height: 100%; justify-content: space-between;">
<div class="overview-metric">
<div class="overview-metric-num" style="color: #f1f5f9;">{_n_total}</div>
<div class="overview-metric-lbl">Evaluated</div>
</div>
<div class="overview-metric">
<div class="overview-metric-num" style="color: #059669;">{_n_shortlisted}</div>
<div class="overview-metric-lbl">Shortlisted</div>
</div>
<div class="overview-metric">
<div class="overview-metric-num" style="color: #4f46e5;">{_avg_score}</div>
<div class="overview-metric-lbl">Average Score</div>
</div>
<div class="overview-metric">
<div class="overview-metric-num" style="color: #0891b2;">{_highest_score}</div>
<div class="overview-metric-lbl">Highest Score</div>
</div>
</div>
''',
                unsafe_allow_html=True
            )
            
    st.markdown("<br>", unsafe_allow_html=True)
    
    st.markdown('<div class="section-heading">Pipeline Overview</div>', unsafe_allow_html=True)
    _pov_cols = st.columns(4, gap="medium")
    _pov_data = [
        (str(_n_total),       "Candidates Evaluated", "#6366f1"),
        (str(_n_shortlisted), "Shortlisted",          "#10b981"),
        (str(_n_hold),        "On Hold",              "#f59e0b"),
        (str(_n_rejected),    "Rejected",             "#ef4444"),
    ]
    for _pc, (_pnum, _plbl, _pcol) in zip(_pov_cols, _pov_data):
        _pc.markdown(
            f'<div class="pipeline-stat-card" style="border-top:3px solid {_pcol}40;">'
            f'<div class="pipeline-stat-num" style="color:{_pcol};">{_pnum}</div>'
            f'<div class="pipeline-stat-lbl">{_plbl}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Filter ───────────────────────────────────────────────────────────────
    _filter_options = ["All", "Shortlisted", "On Hold", "Rejected", "Recommended Only", "High Confidence Only"]
    _fc1, _fc2 = st.columns([0.58, 0.42])
    with _fc1:
        _filter_sel = st.radio(
            "Filter Candidates",
            options=_filter_options,
            index=_filter_options.index(st.session_state.get("pipeline_filter", "All")),
            horizontal=True,
            key="pipeline_filter_radio",
            label_visibility="collapsed",
        )
        st.session_state.pipeline_filter = _filter_sel
    with _fc2:
        _filter_desc_map = {
            "All":                  "Showing all evaluated candidates.",
            "Shortlisted":          f"Showing {_n_shortlisted} shortlisted candidate(s).",
            "On Hold":              f"Showing {_n_hold} candidate(s) on hold.",
            "Rejected":             f"Showing {_n_rejected} rejected candidate(s).",
            "Recommended Only":     "Candidate Fit Score >= 65.",
            "High Confidence Only": "AI Confidence = High.",
        }
        st.markdown(
            f'<div style="padding:9px 14px;font-size:0.78rem;color:#64748b;margin-top:2px;">'
            f'Filter: {_filter_desc_map.get(_filter_sel, "")}</div>',
            unsafe_allow_html=True,
        )

    # Apply filter ─────────────────────────────────────────────────────────────
    if _filter_sel == "Shortlisted":
        _filtered_ranked = [e for e in _ranked if _get_status(e) == "shortlisted"]
    elif _filter_sel == "On Hold":
        _filtered_ranked = [e for e in _ranked if _get_status(e) == "hold"]
    elif _filter_sel == "Rejected":
        _filtered_ranked = [e for e in _ranked if _get_status(e) == "rejected"]
    elif _filter_sel == "Recommended Only":
        _filtered_ranked = [e for e in _ranked if (_active_score(e) or 0) >= 65]
    elif _filter_sel == "High Confidence Only":
        _filtered_ranked = [e for e in _ranked if (e["result"].confidence_level or "").lower() == "high"]
    else:
        _filtered_ranked = _ranked

    if not _filtered_ranked and _filter_sel != "All":
        st.markdown(
            f'<div class="info-box" style="margin:10px 0 16px;">No candidates match '
            f'the <strong>{_filter_sel}</strong> filter yet. '
            f'Use the action buttons in each report below to set candidate statuses.</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<span class="section-anchor" id="leaderboard"></span>', unsafe_allow_html=True)

    # ── Candidate Leaderboard ────────────────────────────────────────────────
    if len(_filtered_ranked) > 1:
        st.markdown('<div class="section-heading">Candidate Leaderboard</div>', unsafe_allow_html=True)

        # ── Ranking Mode Selector ─────────────────────────────────────────
        _mode_descriptions = {
            "Combined Score":          "Ranks using both ATS & HirangAI metrics  (Fit × 70% + ATS × 30%)",
            "Candidate Fit Score Only":   "Ranks using AI evaluation and HR weighting only",
            "ATS Score Only":          "Ranks using ATS Compatibility only",
        }
        _rm_col1, _rm_col2 = st.columns([0.42, 0.58], gap="medium")
        with _rm_col1:
            _mode_sel = st.radio(
                label="Ranking Mode",
                options=["Combined Score", "Candidate Fit Score Only", "ATS Score Only"],
                index=["Combined Score", "Candidate Fit Score Only", "ATS Score Only"].index(
                    st.session_state.ranking_mode
                ),
                key="ranking_mode_radio",
                horizontal=True,
                help="Choose how candidates are ordered in the leaderboard.",
            )
            if _mode_sel != st.session_state.ranking_mode:
                st.session_state.ranking_mode = _mode_sel
                st.rerun()
        with _rm_col2:
            st.markdown(
                f'<div style="background:rgba(99,102,241,0.07);border:1px solid rgba(99,102,241,0.2);'
                f'border-radius:10px;padding:10px 14px;font-size:0.8rem;color:#94a3b8;line-height:1.5;'
                f'margin-top:6px;">'
                f'<strong style="color:#818cf8;">{_mode_sel}</strong> — '
                f'{_mode_descriptions.get(_mode_sel, "")}'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.markdown("<br>", unsafe_allow_html=True)

        _rank_medal  = {1: "🥇", 2: "🥈", 3: "🥉"}
        _rank_border = {1: "#f59e0b", 2: "#94a3b8", 3: "#cd7f32"}
        _rank_bg     = {1: "rgba(245,158,11,0.08)", 2: "rgba(148,163,184,0.06)", 3: "rgba(205,127,50,0.06)"}

        _rec_colours = {
            "Exceptional Candidate":      "#10b981",
            "Outstanding Match":          "#10b981",
            "Strong Hire":                "#6366f1",
            "Recommended":                "#6366f1",
            "Potential Fit":              "#f59e0b",
            "Consider with Reservations": "#f59e0b",
            "Weak Match":                 "#ef4444",
            "Not Recommended":            "#ef4444",
        }

        from candidate_evaluator import _score_to_recommendation as _s2r

        # Header row
        # SaaS Clean Table columns
        _lh = st.columns([0.06, 0.24, 0.12, 0.12, 0.12, 0.20, 0.14])
        _lh_labels = ["Rank", "Candidate", "Final Score", "Fit Score", "ATS Score", "Recommendation", "Status"]
        _lh_tips   = ["", "Name & Resume file", f"Combined: Fit×{int(_TL_WEIGHT*100)}% + ATS×{int(_ATS_WEIGHT*100)}%", "Candidate Fit Score", "ATS Compatibility", "", "Review status"]
        for _col, _txt, _tip in zip(_lh, _lh_labels, _lh_tips):
            _col.markdown(
                f'<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;'
                f'text-transform:uppercase;color:#64748b;padding:6px 0 4px;"'
                f'{f" title=\"{_tip}\"" if _tip else ""}>{_txt}</div>',
                unsafe_allow_html=True,
            )

        for _rank, _entry in enumerate(_filtered_ranked, 1):
            _r       = _entry["result"]
            _sc      = _active_score(_entry)
            _rec     = _s2r(_sc) if _sc is not None else (_r.recommendation or "—")
            _conf    = _r.confidence_level or "Medium"
            _name    = _r.candidate_name or _entry["file"].replace(".pdf", "")
            
            # ATS score from this entry
            _ats_r   = getattr(_r, "ats_result", None)
            _ats_s   = _ats_r.ats_score if _ats_r else None
            _ats_lbl_col = (
                "#059669" if (_ats_s or 0) >= 80 else
                "#d97706" if (_ats_s or 0) >= 60 else "#dc2626"
            )

            # Final / combined score
            _fin_s   = _combined_score(_entry)
            _fin_col = (
                "#059669" if (_fin_s or 0) >= 85 else
                "#4f46e5" if (_fin_s or 0) >= 65 else
                "#d97706" if (_fin_s or 0) >= 40 else "#dc2626"
            )
            _sc_col  = (
                "#059669" if (_sc or 0) >= 85 else
                "#4f46e5" if (_sc or 0) >= 65 else
                "#d97706" if (_sc or 0) >= 40 else "#dc2626"
            )

            _lb_stat     = _get_status(_entry)
            _lb_stat_cls = {"shortlisted": "status-shortlisted", "hold": "status-hold", "rejected": "status-rejected"}.get(_lb_stat, "status-new")
            _lb_stat_lbl = {"shortlisted": "Shortlisted", "hold": "On Hold", "rejected": "Rejected"}.get(_lb_stat, "New")

            _row = st.columns([0.06, 0.24, 0.12, 0.12, 0.12, 0.20, 0.14])
            
            # Clean uniform row rendering (no big borders/medal cells)
            _cell_style = "font-size:0.85rem; padding:8px 0; color:#e2e8f0; display:flex; align-items:center;"
            
            # Rank
            _row[0].markdown(f'<div style="{_cell_style}font-weight:700;">#{_rank}</div>', unsafe_allow_html=True)
            # Candidate
            _row[1].markdown(
                f'<div style="{_cell_style}flex-direction:column;align-items:flex-start;line-height:1.4;">'
                f'<span style="font-weight:600;color:#f1f5f9;">{_name}</span>'
                f'<span style="font-size:0.7rem;color:#475569;">{_entry["file"]}</span></div>',
                unsafe_allow_html=True,
            )
            # Final Score
            _row[2].markdown(f'<div style="{_cell_style}font-weight:800;color:{_fin_col};font-size:1rem;">{_fin_s if _fin_s is not None else "—"}</div>', unsafe_allow_html=True)
            # Fit Score
            _row[3].markdown(f'<div style="{_cell_style}font-weight:600;color:{_sc_col};">{_sc if _sc is not None else "N/A"}</div>', unsafe_allow_html=True)
            # ATS Score
            _row[4].markdown(f'<div style="{_cell_style}color:{_ats_lbl_col};">{_ats_s if _ats_s is not None else "—"}</div>', unsafe_allow_html=True)
            # Recommendation
            _row[5].markdown(f'<div style="{_cell_style}font-weight:500;">{_rec}</div>', unsafe_allow_html=True)
            # Status
            _row[6].markdown(f'<div style="{_cell_style}"><span class="status-badge {_lb_stat_cls}">{_lb_stat_lbl}</span></div>', unsafe_allow_html=True)
            
        st.markdown("<hr style='margin:12px 0;'>", unsafe_allow_html=True)
        # ── Ranking mode legend ───────────────────────────────────────────────
        _rm_now = st.session_state.get("ranking_mode", "Combined Score")
        if _rm_now == "Combined Score":
            _legend_txt = (
                f"<strong>Final Score</strong> = HirangAI Fit × {int(_TL_WEIGHT*100)}% + ATS × {int(_ATS_WEIGHT*100)}%. "
                "The <strong>Final</strong> column drives the current ranking."
            )
        elif _rm_now == "Candidate Fit Score Only":
            _legend_txt = "Ranking driven by <strong>Candidate Fit Score</strong> — AI evaluation weighted by HR configuration."
        else:
            _legend_txt = "Ranking driven by <strong>ATS Compatibility Score</strong> — keyword and structure matching only."
        st.markdown(
            f'<div class="info-box" style="font-size:0.78rem;margin-top:4px;">{_legend_txt}</div>',
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)

    # ── Shortlist Dashboard ─────────────────────────────────────────────
    _shortlisted_all = sorted(
        [e for e in _ranked if _get_status(e) == "shortlisted"],
        key=lambda e: (_ranking_score(e) or 0),
        reverse=True,
    )
    if _shortlisted_all:
        st.markdown('<div class="section-heading">Shortlisted Candidates</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="info-box" style="margin-bottom:12px;font-size:0.82rem;">'
            f'<strong>{len(_shortlisted_all)}</strong> candidate(s) shortlisted for interview consideration. '
            f'Sorted by current ranking score.</div>',
            unsafe_allow_html=True,
        )
        for _si, _se in enumerate(_shortlisted_all, 1):
            _sr      = _se["result"]
            _sname   = _sr.candidate_name or _se["file"].replace(".pdf", "")
            _stl     = _active_score(_se)
            _sats_r  = getattr(_sr, "ats_result", None)
            _sats_v  = _sats_r.ats_score if _sats_r else None
            _sfin    = _combined_score(_se)
            _srec    = _sr.recommendation or "—"
            _sconf   = _sr.confidence_level or "—"
            _sirec_l, _sirec_c = _interview_rec(_sfin)
            _stl_c   = "#10b981" if (_stl or 0) >= 85 else "#6366f1" if (_stl or 0) >= 65 else "#f59e0b"
            _sfin_c  = "#10b981" if (_sfin or 0) >= 85 else "#818cf8" if (_sfin or 0) >= 65 else "#f59e0b"
            st.markdown(
                f'<div class="shortlist-cand-card">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;">'
                f'<div>'
                f'<span style="font-size:0.7rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#10b981;">#{_si}</span>&nbsp;'
                f'<span style="font-size:0.95rem;font-weight:700;color:#e2e8f0;">{_sname}</span>'
                f'</div>'
                f'<div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center;">'
                f'<span style="font-size:0.78rem;font-weight:700;color:{_stl_c};">Fit&nbsp;{_stl if _stl is not None else "N/A"}</span>'
                f'<span style="color:#334155;">|</span>'
                f'<span style="font-size:0.78rem;font-weight:700;color:#06b6d4;">ATS&nbsp;{_sats_v if _sats_v is not None else "N/A"}</span>'
                f'<span style="color:#334155;">|</span>'
                f'<span style="font-size:0.78rem;font-weight:700;color:{_sfin_c};">Final&nbsp;{_sfin if _sfin is not None else "N/A"}</span>'
                f'<span style="color:#334155;">|</span>'
                f'<span style="font-size:0.78rem;color:#94a3b8;">{_srec}</span>'
                f'<span style="color:#334155;">|</span>'
                f'<span style="font-size:0.78rem;font-weight:600;color:{_sirec_c};">'
                f'{_sirec_l}</span>'
                f'</div></div></div>',
                unsafe_allow_html=True,
            )
        st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<span class="section-anchor" id="evaluations"></span>', unsafe_allow_html=True)

    # ── Individual evaluation reports ──────────────────────────────────────────
    st.markdown('<div class="section-heading">Individual Evaluation Reports</div>', unsafe_allow_html=True)



    _tab_labels = [
        f"#{_rank_idx} — {_entry['result'].candidate_name or _entry['file'].replace('.pdf','')}"
        if len(_ranked) > 1
        else _entry['result'].candidate_name or _entry['file'].replace('.pdf','')
        for _rank_idx, _entry in enumerate(_filtered_ranked, 1)
    ]
    _cand_tabs = st.tabs(_tab_labels) if _filtered_ranked else []

    for _rank_idx, (_entry, _tab) in enumerate(zip(_filtered_ranked, _cand_tabs), 1):
        with _tab:
            result      = _entry["result"]
            result_file = _entry["file"]
            # Render errors
            if result.errors:
                for err in result.errors:
                    st.markdown(
                        f'<div class="warning-box">⚠️ <strong>Error:</strong> {err}</div>',
                        unsafe_allow_html=True,
                    )

            if not result.errors and result.recommendation is not None:
                # Terminology & score prep
                _curr_stat     = _get_status(_entry)
                _stat_lbl_map  = {"shortlisted": "Shortlisted", "hold": "On Hold", "rejected": "Rejected"}
                _stat_cls_map  = {"shortlisted": "status-shortlisted", "hold": "status-hold", "rejected": "status-rejected"}
                _curr_stat_txt = _stat_lbl_map.get(_curr_stat, "Not Reviewed")
                _curr_stat_cls = _stat_cls_map.get(_curr_stat, "status-new")

                candidate_label = result.candidate_name or result_file.replace(".pdf", "")
                ai_baseline     = result.match_score

                from prompts import CATEGORY_KEYS, CATEGORY_LABELS
                _hr_w     = st.session_state.hr_weights
                _hr_total = sum(_hr_w.values())
                _has_cats = hasattr(result, "category_scores") and bool(result.category_scores)

                if _has_cats and _hr_total == 100:
                    _weighted_sum = sum(
                        result.category_scores.get(k, 0) * (_hr_w.get(k, 0) / 100)
                        for k in CATEGORY_KEYS
                    )
                    score = max(0, min(100, round(_weighted_sum)))
                else:
                    score = ai_baseline

                from candidate_evaluator import _score_to_recommendation as _s2r
                active_recommendation = _s2r(score) if score is not None else result.recommendation

                if score is not None:
                    if score >= 85:
                        score_colour, score_bg = "#059669", "rgba(5,150,105,0.05)"
                    elif score >= 65:
                        score_colour, score_bg = "#4f46e5", "rgba(79,70,229,0.05)"
                    elif score >= 40:
                        score_colour, score_bg = "#d97706", "rgba(217,119,6,0.05)"
                    else:
                        score_colour, score_bg = "#dc2626", "rgba(220,38,38,0.05)"
                else:
                    score_colour, score_bg = "#4f46e5", "rgba(79,70,229,0.03)"

                rec_colours = {
                    "Exceptional Candidate":      ("#059669", "rgba(5,150,105,0.06)"),
                    "Outstanding Match":          ("#059669", "rgba(5,150,105,0.05)"),
                    "Strong Hire":                ("#4f46e5", "rgba(79,70,229,0.06)"),
                    "Recommended":                ("#4f46e5", "rgba(79,70,229,0.05)"),
                    "Potential Fit":              ("#d97706", "rgba(217,119,6,0.06)"),
                    "Consider with Reservations": ("#d97706", "rgba(217,119,6,0.05)"),
                    "Weak Match":                 ("#dc2626", "rgba(220,38,38,0.05)"),
                    "Not Recommended":            ("#dc2626", "rgba(220,38,38,0.06)"),
                    "Hire":    ("#4f46e5", "rgba(79,70,229,0.06)"),
                    "Maybe":   ("#d97706", "rgba(217,119,6,0.06)"),
                    "No Hire": ("#dc2626", "rgba(220,38,38,0.06)"),
                }
                rec_fg, rec_bg = rec_colours.get(active_recommendation, ("#64748b", "rgba(100,116,139,0.05)"))

                conf_colours = {
                    "High":   ("#059669", "rgba(5,150,105,0.04)"),
                    "Medium": ("#d97706", "rgba(217,119,6,0.04)"),
                    "Low":    ("#dc2626", "rgba(220,38,38,0.04)"),
                }
                conf_fg, conf_bg = conf_colours.get(result.confidence_level or "Medium", ("#64748b", "rgba(100,116,139,0.04)"))

                _ats_res = result.ats_result
                _ats_val = _ats_res.ats_score if _ats_res else None
                if _ats_val is not None:
                    if _ats_val >= 90:
                        _ats_c, _ats_bg = "#0891b2", "rgba(8,145,178,0.05)"
                    elif _ats_val >= 80:
                        _ats_c, _ats_bg = "#0891b2", "rgba(8,145,178,0.04)"
                    elif _ats_val >= 70:
                        _ats_c, _ats_bg = "#06b6d4", "rgba(6,182,212,0.04)"
                    elif _ats_val >= 60:
                        _ats_c, _ats_bg = "#d97706", "rgba(217,119,6,0.04)"
                    else:
                        _ats_c, _ats_bg = "#dc2626", "rgba(220,38,38,0.04)"
                else:
                    _ats_c, _ats_bg = "#64748b", "rgba(100,116,139,0.03)"

                _fin_irec = _combined_score(_entry)
                _irec_lbl_r, _irec_col_r = _interview_rec(_fin_irec)

                # Initialize Tabs
                tab_overview, tab_ats, tab_fit, tab_strengths, tab_gaps = st.tabs([
                    "Overview", "ATS Analysis", "Candidate Fit", "Strengths", "Skill Gaps"
                ])

                # ── OVERVIEW TAB ─────────────────────────────────────────────
                with tab_overview:
                    st.markdown("<br>", unsafe_allow_html=True)
                    # Pipeline Status Actions
                    _sa1, _sa2, _sa3, _sa_info = st.columns([0.15, 0.13, 0.12, 0.60])
                    with _sa1:
                        if st.button("Shortlist", key=f"btn_sl_{_rank_idx}", type="primary" if _curr_stat == "shortlisted" else "secondary", use_container_width=True):
                            st.session_state.candidate_statuses[result_file] = "shortlisted"
                            st.rerun()
                    with _sa2:
                        if st.button("Hold", key=f"btn_ho_{_rank_idx}", type="primary" if _curr_stat == "hold" else "secondary", use_container_width=True):
                            st.session_state.candidate_statuses[result_file] = "hold"
                            st.rerun()
                    with _sa3:
                        if st.button("Reject", key=f"btn_re_{_rank_idx}", type="primary" if _curr_stat == "rejected" else "secondary", use_container_width=True):
                            st.session_state.candidate_statuses[result_file] = "rejected"
                            st.rerun()
                    with _sa_info:
                        st.markdown(f'<div style="padding: 6px 0;">Status: <span class="status-badge {_curr_stat_cls}">{_curr_stat_txt}</span></div>', unsafe_allow_html=True)

                    st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)

                    # Metrics Row
                    h1, h2, h3, h4 = st.columns(4, gap="medium")
                    with h1:
                        if _ats_val is not None:
                            _ats_lbl = _ats_res.ats_label if _ats_res else ""
                            st.markdown(
                                f'''
<div style="background:{_ats_bg}; border:1px solid {_ats_c}25; border-radius:6px; padding:16px 14px; text-align:center;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:{_ats_c};margin-bottom:4px;">ATS Score</div>
<div style="font-size:2rem;font-weight:800;color:{_ats_c};line-height:1;">{_ats_val}<span style="font-size:0.9rem;opacity:0.7;">/100</span></div>
<div style="font-size:0.68rem;color:{_ats_c}d0;margin-top:4px;font-weight:600;">{_ats_lbl}</div>
</div>
''', unsafe_allow_html=True
                            )
                        else:
                            st.markdown('<div style="background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.05);border-radius:6px;padding:16px 14px;text-align:center;"><div style="font-size:0.65rem;color:#64748b;text-transform:uppercase;">ATS Score</div><div style="font-size:1.4rem;font-weight:700;color:#475569;margin-top:4px;">N/A</div></div>', unsafe_allow_html=True)
                    with h2:
                        if score is not None:
                            st.markdown(
                                f'''
<div style="background:{score_bg}; border:1px solid {score_colour}25; border-radius:6px; padding:16px 14px; text-align:center;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:{score_colour};margin-bottom:4px;">Fit Score</div>
<div style="font-size:2rem;font-weight:800;color:{score_colour};line-height:1;">{score}<span style="font-size:0.9rem;opacity:0.7;">/100</span></div>
</div>
''', unsafe_allow_html=True
                            )
                        else:
                            st.markdown('<div style="background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.05);border-radius:6px;padding:16px 14px;text-align:center;"><div style="font-size:0.65rem;color:#64748b;text-transform:uppercase;">Fit Score</div><div style="font-size:1.4rem;font-weight:700;color:#475569;margin-top:4px;">N/A</div></div>', unsafe_allow_html=True)
                    with h3:
                        st.markdown(
                            f'''
<div style="background:{rec_bg}; border:1px solid {rec_fg}25; border-radius:6px; padding:16px 14px; text-align:center; height:100%; display:flex; flex-direction:column; justify-content:center;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:{rec_fg};margin-bottom:4px;">Recommendation</div>
<div style="font-size:0.95rem;font-weight:700;color:{rec_fg};">{active_recommendation}</div>
</div>
''', unsafe_allow_html=True
                        )
                    with h4:
                        st.markdown(
                            f'''
<div style="background:{conf_bg}; border:1px solid {conf_fg}25; border-radius:6px; padding:16px 14px; text-align:center;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:{conf_fg};margin-bottom:4px;">Confidence</div>
<div style="font-size:1.3rem;font-weight:700;color:{conf_fg};margin-top:2px;">{result.confidence_level or 'Medium'}</div>
</div>
''', unsafe_allow_html=True
                        )

                    # Interview Recommendation
                    if _fin_irec is not None:
                        st.markdown(
                            f'<div class="interview-rec-strip" style="background:{_irec_col_r}08;border:1px solid {_irec_col_r}20;border-left:3px solid {_irec_col_r};margin-top:14px;">'
                            f'<span style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:{_irec_col_r};">Interview Recommendation</span>'
                            f'<span style="font-size:0.95rem;font-weight:700;color:{_irec_col_r};margin-left:10px;">{_irec_lbl_r}</span>'
                            f'<span style="font-size:0.75rem;color:#64748b;margin-left:auto;">Final Score: {_fin_irec}/100</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                    # Score explanation cards
                    st.markdown("<div class='section-heading'>Evaluation Summary</div>", unsafe_allow_html=True)
                    _evs_left, _evs_right = st.columns(2, gap="large")
                    
                    def _ats_explanation(ats_v, ats_lbl, missing_kws):
                        if ats_v is None: return "ATS score could not be computed."
                        missing_count = len(missing_kws)
                        if ats_v >= 90: return f"This candidate's resume is an excellent keyword and structural match for the role. Almost all required skills and sections are present."
                        elif ats_v >= 80: return f"Strong ATS alignment — the resume covers the vast majority of required skills. " + (f"{missing_count} preferred keyword(s) were not detected." if missing_count else "")
                        elif ats_v >= 70: return f"Good overall keyword match with a few gaps. " + (f"Missing: {', '.join(missing_kws[:4])}." if missing_kws else "")
                        elif ats_v >= 60: return f"Moderate ATS alignment. Several required or preferred keywords are absent. " + (f"Missing: {', '.join(missing_kws[:4])}." if missing_kws else "")
                        else: return f"Low ATS compatibility — the resume lacks many keywords expected by standard screening systems. " + (f"Missing: {', '.join(missing_kws[:5])}." if missing_kws else "")

                    def _tl_explanation(tl_v, rec_label, ats_v):
                        if tl_v is None: return "HirangAI score could not be computed."
                        gap = (tl_v - (ats_v or tl_v))
                        if abs(gap) <= 10: txt = f"The HirangAI score of {tl_v}/100 is closely aligned with the ATS score."
                        elif gap > 10: txt = f"Despite lower ATS keyword match, the candidate scores {tl_v}/100 on HirangAI, reflecting strong demonstrated skills, relevant experience, and role suitability."
                        else: txt = f"The candidate's resume keyword coverage outpaces the overall AI-assessed fit ({tl_v}/100)."
                        return f"{txt} Recommendation: **{rec_label}**."

                    _eval_ats_val    = _ats_val
                    _eval_ats_lbl    = (_ats_res.ats_label if _ats_res else "N/A")
                    _eval_missing    = list((_ats_res.required_keywords_missing if _ats_res else []) + (_ats_res.preferred_keywords_missing if _ats_res else []))
                    _eval_tl         = score
                    _eval_fin        = _combined_score(_entry)

                    with _evs_left:
                        st.markdown(
                            f'''
<div style="background:rgba(6,182,212,0.03);border:1px solid rgba(6,182,212,0.12);border-left:3px solid #06b6d4;border-radius:6px;padding:14px;height:100%;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#06b6d4;margin-bottom:6px;">ATS Compatibility Summary</div>
<div style="font-size:1.3rem;font-weight:800;color:#06b6d4;margin-bottom:4px;">{_eval_ats_val if _eval_ats_val is not None else 'N/A'}<span style="font-size:0.75rem;opacity:0.7;">/100</span></div>
<div style="font-size:0.8rem;color:#cbd5e1;line-height:1.5;">{_ats_explanation(_eval_ats_val, _eval_ats_lbl, _eval_missing)}</div>
</div>
''', unsafe_allow_html=True
                        )
                    with _evs_right:
                        st.markdown(
                            f'''
<div style="background:rgba(79,70,229,0.03);border:1px solid rgba(79,70,229,0.12);border-left:3px solid #4f46e5;border-radius:6px;padding:14px;height:100%;">
<div style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#818cf8;margin-bottom:6px;">HirangAI Fit Summary</div>
<div style="font-size:1.3rem;font-weight:800;color:{score_colour};margin-bottom:4px;">{_eval_tl if _eval_tl is not None else 'N/A'}<span style="font-size:0.75rem;opacity:0.7;">/100</span></div>
<div style="font-size:0.8rem;color:#cbd5e1;line-height:1.5;">{_tl_explanation(_eval_tl, active_recommendation, _eval_ats_val)}</div>
</div>
''', unsafe_allow_html=True
                        )

                    if _eval_fin is not None:
                        _fin_c2 = "#059669" if _eval_fin >= 85 else "#4f46e5" if _eval_fin >= 65 else "#d97706" if _eval_fin >= 40 else "#dc2626"
                        st.markdown(
                            f'<div style="background:rgba(79,70,229,0.04);border:1px solid rgba(79,70,229,0.15);border-radius:6px;padding:10px 14px;margin-top:12px;display:flex;align-items:center;gap:12px;">'
                            f'<span style="font-size:0.65rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#818cf8;">Final Ranking Score</span>'
                            f'<span style="font-size:1.3rem;font-weight:800;color:{_fin_c2};">{_eval_fin}</span>'
                            f'<span style="font-size:0.75rem;opacity:0.6;color:#94a3b8;">/100</span>'
                            f'<span style="font-size:0.75rem;color:#64748b;margin-left:auto;">Fit {_eval_tl} × {int(_TL_WEIGHT*100)}% + ATS {_eval_ats_val} × {int(_ATS_WEIGHT*100)}%</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                # ── ATS ANALYSIS TAB ─────────────────────────────────────────
                with tab_ats:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if _ats_res is not None:
                        # Sub-score bars
                        st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#06b6d4;margin-bottom:10px;">ATS Sub-Scores</div>', unsafe_allow_html=True)
                        _sub_col1, _sub_col2 = st.columns(2, gap="large")
                        _sub_items = list((_ats_res.sub_scores or {}).items())
                        for _si, (_slabel, _sval) in enumerate(_sub_items):
                            _scol = _sub_col1 if _si % 2 == 0 else _sub_col2
                            _sval_c = "#059669" if _sval >= 80 else "#06b6d4" if _sval >= 60 else "#d97706" if _sval >= 40 else "#dc2626"
                            _scol.markdown(
                                f'''
                                <div style="margin-bottom:10px;">
                                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
                                        <span style="font-size:0.78rem;font-weight:600;color:#cbd5e1;">{_slabel}</span>
                                        <span style="font-size:0.8rem;font-weight:700;color:{_sval_c};">{_sval}</span>
                                    </div>
                                    <div class="ats-bar-track">
                                        <div class="ats-bar-fill" style="--bar-w:{_sval}%;background:{_sval_c};"></div>
                                    </div>
                                </div>
                                ''', unsafe_allow_html=True
                            )

                        st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)
                        _kw_left, _kw_right = st.columns(2, gap="large")

                        with _kw_left:
                            st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#059669;margin-bottom:6px;">Matched Keywords</div>', unsafe_allow_html=True)
                            _matched_all = _ats_res.required_keywords_matched + _ats_res.preferred_keywords_matched
                            if _matched_all:
                                _chip_html = "".join(f'<span class="ats-chip ats-chip-match">{kw}</span>' for kw in _matched_all[:30])
                                st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:2px;">{_chip_html}</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="color:#64748b;font-size:0.8rem;">No keywords matched.</div>', unsafe_allow_html=True)

                        with _kw_right:
                            st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#d97706;margin-bottom:6px;">Missing Keywords</div>', unsafe_allow_html=True)
                            _missing_all = _ats_res.required_keywords_missing + _ats_res.preferred_keywords_missing
                            if _missing_all:
                                _miss_html = "".join(f'<span class="ats-chip ats-chip-miss">{kw}</span>' for kw in _missing_all[:30])
                                st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:2px;">{_miss_html}</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="color:#059669;font-size:0.8rem;">All keywords matched!</div>', unsafe_allow_html=True)

                        st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)
                        st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#06b6d4;margin-bottom:8px;">Resume Structure</div>', unsafe_allow_html=True)
                        _sect_left, _sect_right = st.columns(2, gap="large")
                        _all_sects = [(s, True) for s in _ats_res.sections_present] + [(s, False) for s in _ats_res.sections_missing]
                        for _si2, (_sname, _spresent) in enumerate(_all_sects):
                            _sc2 = _sect_left if _si2 % 2 == 0 else _sect_right
                            _icon2, _cls2 = ("✓", "ats-section-ok") if _spresent else ("⚠", "ats-section-miss")
                            _sc2.markdown(f'<div class="ats-section-row"><span class="{_cls2}" style="font-weight:700;margin-right:6px;">{_icon2}</span><span style="color:#cbd5e1;">{_sname}</span></div>', unsafe_allow_html=True)

                        # Guide box
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown(
                            '<div style="background:rgba(6,182,212,0.03);border:1px solid rgba(6,182,212,0.15);border-left:3px solid #06b6d4;border-radius:6px;padding:10px 14px;font-size:0.78rem;color:#94a3b8;line-height:1.6;">'
                            '<strong style="color:#06b6d4;">ATS Scoring Rubric</strong><br>'
                            '90–100: Excellent ATS Match | 80–89: Strong | 70–79: Good | 60–69: Moderate | Under 60: Low'
                            '</div>', unsafe_allow_html=True
                        )
                    else:
                        st.markdown('<div style="color:#64748b;font-size:0.85rem;">ATS breakdown not available.</div>', unsafe_allow_html=True)

                # ── CANDIDATE FIT TAB ────────────────────────────────────────
                with tab_fit:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if _has_cats:
                        st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#4f46e5;margin-bottom:12px;">Weighted Categories</div>', unsafe_allow_html=True)
                        cat_left, cat_right = st.columns(2, gap="large")
                        cats = [(key, CATEGORY_LABELS[key], result.category_scores.get(key)) for key in CATEGORY_KEYS]
                        for idx, (key, label, cat_score) in enumerate(cats):
                            col        = cat_left if idx % 2 == 0 else cat_right
                            s          = cat_score if cat_score is not None else 0
                            weight_pct = _hr_w.get(key, 0)
                            
                            if s >= 85: bar_colour, bar_sc = "#059669", "#059669"
                            elif s >= 65: bar_colour, bar_sc = "#4f46e5", "#4f46e5"
                            elif s >= 40: bar_colour, bar_sc = "#d97706", "#d97706"
                            else: bar_colour, bar_sc = "#dc2626", "#dc2626"
                            
                            col.markdown(
                                f'''
                                <div style="margin-bottom:12px;">
                                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
                                        <span style="font-size:0.78rem;font-weight:600;color:#cbd5e1;">{label}</span>
                                        <span style="font-size:0.72rem;color:#64748b;">weight {weight_pct}%</span>
                                    </div>
                                    <div style="display:flex;align-items:center;gap:10px;">
                                        <div class="cat-bar-track">
                                            <div class="cat-bar-fill" style="--bar-w:{s}%;background:{bar_colour};"></div>
                                        </div>
                                        <span style="font-size:0.82rem;font-weight:700;color:{bar_sc};min-width:32px;text-align:right;">{s if cat_score is not None else '—'}</span>
                                    </div>
                                </div>
                                ''', unsafe_allow_html=True
                            )
                        if _hr_total != 100:
                            st.markdown('<div class="warning-box" style="margin-top:10px;">⚠️ Total weights configured must equal 100%.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="color:#64748b;font-size:0.85rem;">Category details not available.</div>', unsafe_allow_html=True)

                # ── STRENGTHS TAB ────────────────────────────────────────────
                with tab_strengths:
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#059669;margin-bottom:8px;">Key Strengths</div>', unsafe_allow_html=True)
                    if result.strengths:
                        bullets = "".join(f'<li style="margin-bottom:5px;">{s}</li>' for s in result.strengths)
                        st.markdown(f'<ul style="color:#94a3b8;font-size:0.82rem;line-height:1.6;padding-left:16px;">{bullets}</ul>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="color:#64748b;font-size:0.8rem;">No strengths found.</div>', unsafe_allow_html=True)

                    if result.experience_summary:
                        st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)
                        st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#4f46e5;margin-bottom:8px;">Experience Summary</div>', unsafe_allow_html=True)
                        st.markdown(f'<div style="color:#cbd5e1;font-size:0.82rem;line-height:1.6;background:rgba(255,255,255,0.01);padding:10px 14px;border:1px solid rgba(255,255,255,0.05);border-radius:6px;">{result.experience_summary}</div>', unsafe_allow_html=True)

                # ── SKILL GAPS TAB ───────────────────────────────────────────
                with tab_gaps:
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown('<div style="font-size:0.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#d97706;margin-bottom:8px;">Detected Skill Gaps</div>', unsafe_allow_html=True)
                    if result.skill_gaps:
                        gaps = "".join(f'<li style="margin-bottom:5px;">{g}</li>' for g in result.skill_gaps)
                        st.markdown(f'<ul style="color:#94a3b8;font-size:0.82rem;line-height:1.6;padding-left:16px;">{gaps}</ul>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="color:#059669;font-size:0.8rem;">No skill gaps detected.</div>', unsafe_allow_html=True)

                    if result.raw_llm_response:
                        st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)
                        with st.expander("Full AI Analysis Details", expanded=False):
                            st.markdown(result.raw_llm_response, unsafe_allow_html=True)

                    st.markdown("<hr style='margin:14px 0;'>", unsafe_allow_html=True)
                    # Report download
                    _ind_pdf = _build_individual_pdf(
                        entry=_entry,
                        active_score=_active_score(_entry),
                        active_recommendation=active_recommendation,
                        hr_weights=_hr_w,
                    )
                    _cand_fname = (result.candidate_name or result_file.replace(".pdf", "")).replace(" ", "_")
                    st.download_button(
                        label="Download PDF Report",
                        data=_ind_pdf,
                        file_name=f"HirangAI_Report_{_cand_fname}.pdf",
                        mime="application/pdf",
                        key=f"dl_ind_pdf_{_rank_idx}",
                        use_container_width=False,
                    )
        st.markdown('<span class="section-anchor" id="comparison"></span>', unsafe_allow_html=True)

# CANDIDATE COMPARISON DASHBOARD
# =============================================================================

if st.session_state.get("batch_results") and len(st.session_state.batch_results) >= 2:
    from prompts import CATEGORY_KEYS, CATEGORY_LABELS
    from candidate_evaluator import _score_to_recommendation as _s2r

    _br2      = st.session_state.batch_results
    _hr_w2    = st.session_state.hr_weights
    _hr_tot2  = sum(_hr_w2.values())

    def _active_score2(entry: dict) -> int | None:
        r = entry["result"]
        if hasattr(r, "category_scores") and r.category_scores and _hr_tot2 == 100:
            ws = sum(r.category_scores.get(k, 0) * (_hr_w2.get(k, 0) / 100) for k in CATEGORY_KEYS)
            return max(0, min(100, round(ws)))
        return r.match_score

    _ranked2 = sorted(
        _br2,
        key=lambda e: (_active_score2(e) or 0) if not e["result"].errors else -1,
        reverse=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div class="section-heading">Candidate Comparison Dashboard</div>',
        unsafe_allow_html=True,
    )

    _cand_options = [
        (e["result"].candidate_name or e["file"].replace(".pdf", ""))
        for e in _ranked2
    ]

    _selected_names = st.multiselect(
        "Select 2–4 candidates to compare side-by-side:",
        options=_cand_options,
        default=_cand_options[:min(2, len(_cand_options))],
        max_selections=4,
        key="comparison_multiselect",
        help="Choose candidates from your evaluated pool to see a side-by-side breakdown.",
    )

    if len(_selected_names) < 2:
        st.markdown(
            '<div class="info-box">ℹ️ Select at least <strong>2 candidates</strong> to enable the comparison view.</div>',
            unsafe_allow_html=True,
        )
    else:
        # Filter entries to only selected candidates (preserve rank order)
        _compare_entries = [
            e for e in _ranked2
            if (e["result"].candidate_name or e["file"].replace(".pdf", "")) in _selected_names
        ]

        _n_cols = len(_compare_entries)
        _cmp_cols = st.columns(_n_cols, gap="medium")

        _cat_icons2 = {
            "technical_skills":          "",
            "relevant_experience":       "",
            "education":                 "",
            "certifications":            "",
            "communication_soft_skills": "",
        }

        for _col_ui, _entry in zip(_cmp_cols, _compare_entries):
            _r2     = _entry["result"]
            _sc2    = _active_score2(_entry)
            _rec2   = _s2r(_sc2) if _sc2 is not None else (_r2.recommendation or "—")
            _cname2 = _r2.candidate_name or _entry["file"].replace(".pdf", "")

            _sc_col2 = (
                "#10b981" if (_sc2 or 0) >= 85 else
                "#6366f1" if (_sc2 or 0) >= 65 else
                "#f59e0b" if (_sc2 or 0) >= 40 else "#ef4444"
            )
            _rc_map2 = {
                "Exceptional Candidate": "#10b981", "Outstanding Match": "#10b981",
                "Strong Hire": "#6366f1",           "Recommended": "#6366f1",
                "Potential Fit": "#f59e0b",         "Consider with Reservations": "#f59e0b",
                "Weak Match": "#ef4444",            "Not Recommended": "#ef4444",
            }
            _rc_col2 = _rc_map2.get(_rec2, "#94a3b8")

            with _col_ui:
                # ── Name header ───────────────────────────────────────────
                st.markdown(
                    f"""
<div class="glass-card" style="text-align:center;padding:20px 16px 16px;">
<div style="font-size:0.72rem;font-weight:600;letter-spacing:1.2px;
text-transform:uppercase;color:#6366f1;margin-bottom:6px;">
Candidate
</div>
<div style="font-size:1rem;font-weight:700;color:#e2e8f0;
word-break:break-word;margin-bottom:14px;">
{_cname2}
</div>
<div style="font-size:2.4rem;font-weight:800;color:{_sc_col2};line-height:1;">
{_sc2 if _sc2 is not None else "N/A"}
<span style="font-size:0.9rem;color:{_sc_col2}99;">/100</span>
</div>
<div style="font-size:0.72rem;font-weight:700;color:{_sc_col2};
margin-top:4px;margin-bottom:6px;letter-spacing:0.5px;">
Candidate Fit Score
</div>
<div style="display:inline-block;background:{_rc_col2}22;
border:1px solid {_rc_col2}55;border-radius:100px;
padding:4px 12px;font-size:0.75rem;font-weight:700;
color:{_rc_col2};margin-bottom:10px;">
{_rec2}
</div>
</div>
""",
                    unsafe_allow_html=True,
                )

                # ── ATS Compatibility mini-card in comparison ─────────────────
                _ats_r2  = getattr(_r2, "ats_result", None)
                _ats_s2  = _ats_r2.ats_score if _ats_r2 else None
                _ats_lbl2 = _ats_r2.ats_label if _ats_r2 else "—"
                _ats_c2  = (
                    "#06b6d4" if (_ats_s2 or 0) >= 80 else
                    "#22d3ee" if (_ats_s2 or 0) >= 70 else
                    "#f59e0b" if (_ats_s2 or 0) >= 60 else "#ef4444"
                ) if _ats_s2 is not None else "#64748b"
                st.markdown(
                    f"""
<div style="background:rgba(6,182,212,0.07);border:1px solid {_ats_c2}40;
border-radius:12px;padding:12px 16px;text-align:center;
margin-bottom:10px;">
<div style="font-size:0.68rem;font-weight:700;letter-spacing:1.2px;
text-transform:uppercase;color:{_ats_c2};margin-bottom:4px;">
ATS Compatibility
</div>
<div style="font-size:1.8rem;font-weight:800;color:{_ats_c2};line-height:1;">
{_ats_s2 if _ats_s2 is not None else "N/A"}
<span style="font-size:0.85rem;color:{_ats_c2}99;">/100</span>
</div>
<div style="font-size:0.68rem;color:{_ats_c2}cc;margin-top:3px;font-weight:600;">
{_ats_lbl2}
</div>
</div>
""",
                    unsafe_allow_html=True,
                )

                # ── Category scores ───────────────────────────────────────
                st.markdown(
                    '<div style="font-size:0.72rem;font-weight:700;letter-spacing:1.2px;'
                    'text-transform:uppercase;color:#6366f1;margin:14px 0 8px;">Category Scores</div>',
                    unsafe_allow_html=True,
                )
                _cats2 = _r2.category_scores or {}
                for _key2 in CATEGORY_KEYS:
                    _s2val = _cats2.get(_key2)
                    _s2num = _s2val if _s2val is not None else 0
                    _lbl2  = CATEGORY_LABELS.get(_key2, _key2)
                    _icon2 = _cat_icons2.get(_key2, "•")
                    if _s2num >= 85:
                        _bc2 = "linear-gradient(90deg,#10b981,#34d399)"; _tc2 = "#10b981"
                    elif _s2num >= 65:
                        _bc2 = "linear-gradient(90deg,#6366f1,#818cf8)"; _tc2 = "#6366f1"
                    elif _s2num >= 40:
                        _bc2 = "linear-gradient(90deg,#f59e0b,#fbbf24)"; _tc2 = "#f59e0b"
                    else:
                        _bc2 = "linear-gradient(90deg,#ef4444,#f87171)"; _tc2 = "#ef4444"
                    st.markdown(
                        f"""
<div style="margin-bottom:10px;">
<div style="display:flex;justify-content:space-between;
align-items:center;margin-bottom:4px;">
<span style="font-size:0.75rem;font-weight:600;color:#cbd5e1;">
{_icon2}&nbsp;{_lbl2}
</span>
<span style="font-size:0.78rem;font-weight:700;color:{_tc2};">
{_s2val if _s2val is not None else "—"}
</span>
</div>
<div class="cat-bar-track">
<div class="cat-bar-fill"
style="--bar-w:{_s2num}%;background:{_bc2};">
</div>
</div>
</div>
""",
                        unsafe_allow_html=True,
                    )

                # ── Strengths ─────────────────────────────────────────────
                st.markdown(
                    '<div style="font-size:0.72rem;font-weight:700;letter-spacing:1.2px;'
                    'text-transform:uppercase;color:#6366f1;margin:14px 0 6px;">Key Strengths</div>',
                    unsafe_allow_html=True,
                )
                if _r2.strengths:
                    _str_bullets = "".join(
                        f'<li style="margin-bottom:4px;">{s}</li>'
                        for s in _r2.strengths[:5]
                    )
                    st.markdown(
                        f'<ul style="color:#94a3b8;font-size:0.8rem;line-height:1.6;'
                        f'padding-left:16px;margin:0;">{_str_bullets}</ul>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="color:#475569;font-size:0.8rem;">No strengths extracted.</div>',
                        unsafe_allow_html=True,
                    )

                # ── Skill Gaps ────────────────────────────────────────────
                st.markdown(
                    '<div style="font-size:0.72rem;font-weight:700;letter-spacing:1.2px;'
                    'text-transform:uppercase;color:#6366f1;margin:14px 0 6px;">Skill Gaps</div>',
                    unsafe_allow_html=True,
                )
                if _r2.skill_gaps:
                    _gap_bullets = "".join(
                        f'<li style="margin-bottom:4px;">{g}</li>'
                        for g in _r2.skill_gaps[:5]
                    )
                    st.markdown(
                        f'<ul style="color:#94a3b8;font-size:0.8rem;line-height:1.6;'
                        f'padding-left:16px;margin:0;">{_gap_bullets}</ul>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="color:#475569;font-size:0.8rem;">No skill gaps detected.</div>',
                        unsafe_allow_html=True,
                    )

        st.markdown("<br>", unsafe_allow_html=True)


# =============================================================================
st.markdown('<span class="section-anchor" id="export"></span>', unsafe_allow_html=True)

# RANKINGS EXPORT — PDF + EXCEL
# =============================================================================

if st.session_state.get("batch_results"):
    from prompts import CATEGORY_KEYS
    from candidate_evaluator import _score_to_recommendation as _s2r

    _br_exp    = st.session_state.batch_results
    _hr_w_exp  = st.session_state.hr_weights
    _hr_tot_exp = sum(_hr_w_exp.values())

    def _active_score_exp(entry: dict) -> int | None:
        r = entry["result"]
        if hasattr(r, "category_scores") and r.category_scores and _hr_tot_exp == 100:
            ws = sum(r.category_scores.get(k, 0) * (_hr_w_exp.get(k, 0) / 100) for k in CATEGORY_KEYS)
            return max(0, min(100, round(ws)))
        return r.match_score

    _ranked_exp = sorted(
        _br_exp,
        key=lambda e: (_active_score_exp(e) or 0) if not e["result"].errors else -1,
        reverse=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div class="section-heading">Export Candidate Rankings</div>',
        unsafe_allow_html=True,
    )

    _exp_col1, _exp_col2 = st.columns(2, gap="medium")

    with _exp_col1:
        st.markdown(
            '<div class="glass-card" style="padding:20px 22px 18px;">'
            '<div class="card-label">Rankings PDF</div>'
            '<div class="card-title">Candidate Rankings PDF</div>'
            '<div class="card-desc">Full leaderboard with all scores and category breakdowns.</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        _pdf_bytes = _build_rankings_pdf(
            ranked=_ranked_exp,
            active_score_fn=_active_score_exp,
            template_name=st.session_state.active_template,
        )
        st.download_button(
            label="Download Rankings PDF",
            data=_pdf_bytes,
            file_name="HirangAI_Rankings.pdf",
            mime="application/pdf",
            key="dl_rankings_pdf",
            use_container_width=True,
        )

    with _exp_col2:
        st.markdown(
            '<div class="glass-card" style="padding:20px 22px 18px;">'
            '<div class="card-label">Rankings Excel</div>'
            '<div class="card-title">Candidate Rankings Excel</div>'
            '<div class="card-desc">Spreadsheet with colour-coded scores and full category data.</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        _xlsx_bytes = _build_rankings_excel(
            ranked=_ranked_exp,
            active_score_fn=_active_score_exp,
            template_name=st.session_state.active_template,
        )
        st.download_button(
            label="Download Rankings Excel",
            data=_xlsx_bytes,
            file_name="HirangAI_Rankings.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_rankings_xlsx",
            use_container_width=True,
        )

    # ── Shortlist Export ───────────────────────────────────────────────────
    _statuses_exp = st.session_state.get("candidate_statuses", {})
    _TL_W_EX, _ATS_W_EX = 0.70, 0.30

    def _comb_exp(entry):
        tl  = _active_score_exp(entry)
        _ar = getattr(entry["result"], "ats_result", None)
        ats = _ar.ats_score if _ar else None
        if tl is None and ats is None: return None
        t = tl  if tl  is not None else 0
        a = ats if ats is not None else 0
        return max(0, min(100, round(t * _TL_W_EX + a * _ATS_W_EX)))

    def _irec_label_exp(fs):
        if fs is None: return "—"
        if fs >= 80:   return "Strongly Recommended"
        elif fs >= 65: return "Recommended"
        elif fs >= 50: return "Conditional"
        else:          return "Not Recommended"

    _sl_exp_entries = [
        e for e in _ranked_exp
        if _statuses_exp.get(e["file"]) == "shortlisted"
    ]

    if _sl_exp_entries:
        _sl_sorted = sorted(_sl_exp_entries, key=lambda e: (_comb_exp(e) or 0), reverse=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div class="section-heading">Export Shortlisted Candidates</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div class="info-box" style="margin-bottom:14px;font-size:0.82rem;">'
            f'Exporting <strong>{len(_sl_sorted)}</strong> shortlisted candidate(s).</div>',
            unsafe_allow_html=True,
        )

        _sl_col1, _sl_col2 = st.columns(2, gap="medium")

        with _sl_col1:
            import csv as _csv_mod, io as _io_mod
            _csv_buf = _io_mod.StringIO()
            _csv_w   = _csv_mod.writer(_csv_buf)
            _csv_w.writerow(["Rank", "Candidate", "ATS Score", "Fit Score",
                              "Final Score", "Recommendation", "Confidence",
                              "Interview Recommendation", "Status"])
            for _ci, _ce in enumerate(_sl_sorted, 1):
                _cr    = _ce["result"]
                _ctlv  = _active_score_exp(_ce)
                _catsr = getattr(_cr, "ats_result", None)
                _catsv = _catsr.ats_score if _catsr else None
                _cfin  = _comb_exp(_ce)
                _csv_w.writerow([
                    _ci,
                    _cr.candidate_name or _ce["file"].replace(".pdf", ""),
                    _catsv if _catsv is not None else "N/A",
                    _ctlv  if _ctlv  is not None else "N/A",
                    _cfin  if _cfin  is not None else "N/A",
                    _cr.recommendation or "—",
                    _cr.confidence_level or "—",
                    _irec_label_exp(_cfin),
                    "Shortlisted",
                ])
            st.markdown(
                '<div class="glass-card" style="padding:20px 22px 18px;">'
                '<div class="card-label">Shortlist CSV</div>'
                '<div class="card-title">Shortlisted Candidates CSV</div>'
                '<div class="card-desc">Name, ATS, TL, Final Score, Recommendation, Confidence, Interview Rec.</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                label="Download Shortlist CSV",
                data=_csv_buf.getvalue().encode("utf-8"),
                file_name="HirangAI_Shortlist.csv",
                mime="text/csv",
                key="dl_shortlist_csv",
                use_container_width=True,
            )

        with _sl_col2:
            _sl_pdf = _build_rankings_pdf(
                ranked=_sl_sorted,
                active_score_fn=_active_score_exp,
                template_name=f"Shortlist — {st.session_state.active_template}",
            )
            st.markdown(
                '<div class="glass-card" style="padding:20px 22px 18px;">'
                '<div class="card-label">Shortlist PDF</div>'
                '<div class="card-title">Shortlisted Candidates PDF</div>'
                '<div class="card-desc">Professional PDF report for shortlisted candidates only.</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                label="Download Shortlist PDF",
                data=_sl_pdf,
                file_name="HirangAI_Shortlist.pdf",
                mime="application/pdf",
                key="dl_shortlist_pdf",
                use_container_width=True,
            )
    else:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div class="info-box">ℹNo candidates have been shortlisted yet. '
            'Use the <strong>Shortlist</strong> buttons inside each candidate report '
            'to build your shortlist, then return here to export.</div>',
            unsafe_allow_html=True,
        )

